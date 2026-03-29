import shutil
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

import bot
import onboarding
from excel_manager import ExcelManager


FIXTURE_WORKBOOK = Path(__file__).resolve().parents[1] / "Financial_Tracker.xlsx"


class FakeMessage:
    def __init__(self, text=""):
        self.text = text
        self.replies = []

    async def reply_text(self, text, **kwargs):
        self.replies.append((text, kwargs))


class FakeQuery:
    def __init__(self, data, user_id):
        self.data = data
        self.from_user = SimpleNamespace(id=user_id)
        self.message = FakeMessage()
        self.edits = []

    async def answer(self):
        return None

    async def edit_message_text(self, text, **kwargs):
        self.edits.append((text, kwargs))


class BotReliabilityTestCase(unittest.IsolatedAsyncioTestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "Financial_Tracker.xlsx"
        shutil.copy2(FIXTURE_WORKBOOK, self.workbook_path)
        self.manager = ExcelManager(self.workbook_path)
        self.original_excel = bot.excel
        self.original_is_user_allowed = bot.is_user_allowed
        self.original_mark_complete = onboarding.mark_onboarding_complete
        bot.excel = self.manager
        bot.is_user_allowed = lambda _user_id: True
        onboarding.mark_onboarding_complete = lambda _user_id: None

    def tearDown(self):
        bot.excel = self.original_excel
        bot.is_user_allowed = self.original_is_user_allowed
        onboarding.mark_onboarding_complete = self.original_mark_complete
        self.temp_dir.cleanup()

    async def test_quick_record_accepts_indonesian_style_amounts(self):
        message = FakeMessage("/quick Rp50.000 makan nasi padang")
        update = SimpleNamespace(
            effective_user=SimpleNamespace(id=123),
            effective_message=message,
            message=message,
        )
        context = SimpleNamespace(
            args=["Rp50.000", "makan", "nasi", "padang"],
            user_data={},
            bot_data={"excel_manager": self.manager},
        )

        await bot.quick_record(update, context)

        self.assertTrue(message.replies)
        self.assertIn("Spending recorded", message.replies[-1][0])

        wb = load_workbook(self.workbook_path)
        try:
            ws = wb["Transactions"]
            self.assertEqual(ws["B6"].value, "nasi padang")
            self.assertEqual(ws["C6"].value, "Food & Groceries")
            self.assertEqual(ws["E6"].value, 50000)
        finally:
            wb.close()

    async def test_quick_record_replies_for_unauthorized_user(self):
        bot.is_user_allowed = lambda _user_id: False

        message = FakeMessage("/quick 50000 food lunch")
        update = SimpleNamespace(
            effective_user=SimpleNamespace(id=999),
            effective_message=message,
            message=message,
        )
        context = SimpleNamespace(args=["50000", "food", "lunch"], user_data={}, bot_data={})

        await bot.quick_record(update, context)

        self.assertEqual(message.replies[-1][0], "Unauthorized. Update ALLOWED_USER_IDS to grant access.")

    async def test_onboarding_confirm_writes_all_supported_sheets(self):
        query = FakeQuery("confirm_yes", user_id=123)
        context = SimpleNamespace(
            user_data={
                "onboarding": {
                    "savings": [{"account": "Emergency Fund", "amount": 500000, "goal": 1000000}],
                    "assets": [{"name": "Gold 10g", "value": 12000000, "type": "Gold"}],
                    "income": [{"source": "PT ABC Salary", "amount": 8000000, "frequency": "Monthly"}],
                    "bills": [{"name": "Kos", "amount": 2500000, "category": "Housing"}],
                    "budgets": [{"category": "Housing", "amount": 3000000}],
                }
            },
            bot_data={"excel_manager": self.manager},
        )
        update = SimpleNamespace(callback_query=query)

        result = await onboarding.handle_confirm(update, context)

        self.assertEqual(result, onboarding.ConversationHandler.END)
        self.assertTrue(query.edits)
        self.assertIn("Saving your financial profile", query.edits[0][0])
        self.assertTrue(query.message.replies)
        self.assertIn("SETUP COMPLETE", query.message.replies[-1][0])

        wb = load_workbook(self.workbook_path)
        try:
            self.assertEqual(wb["Savings"]["B7"].value, "Emergency Fund")
            self.assertEqual(wb["Assets"]["B3"].value, "Gold 10g")
            self.assertEqual(wb["Income"]["B4"].value, "PT ABC Salary")
            self.assertEqual(wb["Transactions"]["B6"].value, "Kos (recurring)")
            self.assertEqual(wb["Budget"]["B5"].value, 3000000)
        finally:
            wb.close()


if __name__ == "__main__":
    unittest.main()
