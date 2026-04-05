"""
TDD Tests for Financial Tracker Excel template structure.
Tests verify:
1. Assets sheet has Platform column at position 4
2. Assets sheet has correct 8 columns
3. Debts sheet exists and has correct 11 headers
4. Gain/Loss formula references correct columns
5. Paid % formula references correct columns
"""

import pytest
from pathlib import Path
from openpyxl import load_workbook


@pytest.fixture
def excel_file():
    """Path to the generated Financial_Tracker.xlsx"""
    excel_path = Path(__file__).parent.parent / "Financial_Tracker.xlsx"
    return excel_path


@pytest.fixture
def workbook(excel_file):
    """Load and yield workbook, close after test"""
    wb = load_workbook(excel_file)
    yield wb
    wb.close()


class TestAssetsSheet:
    """Tests for Assets sheet structure"""

    def test_assets_sheet_exists(self, workbook):
        """Verify Assets sheet is present in workbook"""
        assert "Assets" in workbook.sheetnames, "Assets sheet not found"

    def test_assets_has_platform_column(self, workbook):
        """Verify column 4 (D) header is 'Platform'"""
        ws = workbook["Assets"]
        assert ws.cell(row=2, column=4).value == "Platform", (
            f"Column D (4) should be 'Platform', got '{ws.cell(row=2, column=4).value}'"
        )

    def test_assets_column_count(self, workbook):
        """Verify Assets sheet has exactly 8 column headers"""
        ws = workbook["Assets"]
        # Count non-empty headers in row 2
        headers = [ws.cell(row=2, column=col).value for col in range(1, 10)]
        headers = [h for h in headers if h is not None]
        assert len(headers) == 8, f"Expected 8 headers, got {len(headers)}: {headers}"

    def test_assets_column_order(self, workbook):
        """Verify Assets sheet column order is correct"""
        ws = workbook["Assets"]
        expected_headers = [
            "Date Added",
            "Name",
            "Type",
            "Platform",
            "Purchase Value",
            "Current Value",
            "Gain/Loss",
            "Notes",
        ]
        actual_headers = [ws.cell(row=2, column=col).value for col in range(1, 9)]
        assert actual_headers == expected_headers, (
            f"Column order mismatch.\nExpected: {expected_headers}\nGot: {actual_headers}"
        )

    def test_assets_gain_loss_formula(self, workbook):
        """Verify Gain/Loss formula in column 7 references E{row}-F{row}"""
        ws = workbook["Assets"]
        # Check formula in row 3 (first data row)
        formula = ws.cell(row=3, column=7).value
        assert formula is not None, "Column 7 (G) should have a formula"
        assert isinstance(formula, str) and formula.startswith("="), (
            f"Column 7 should contain a formula, got: {formula}"
        )
        # Formula should reference E3-F3 (Purchase Value - Current Value)
        assert "E3" in formula and "F3" in formula, f"Formula should reference E3 and F3, got: {formula}"


class TestDebtsSheet:
    """Tests for Debts sheet structure"""

    def test_debts_sheet_exists(self, workbook):
        """Verify Debts sheet is present in workbook"""
        assert "Debts" in workbook.sheetnames, "Debts sheet not found"

    def test_debts_has_correct_headers(self, workbook):
        """Verify Debts sheet has all 11 correct headers"""
        ws = workbook["Debts"]
        expected_headers = [
            "Date",
            "Name",
            "Type",
            "Bank",
            "Total Loan",
            "Remaining",
            "Monthly Payment",
            "Interest Rate",
            "Tenor (months)",
            "Paid %",
            "Notes",
        ]
        actual_headers = [ws.cell(row=2, column=col).value for col in range(1, 12)]
        assert actual_headers == expected_headers, (
            f"Debts headers mismatch.\nExpected: {expected_headers}\nGot: {actual_headers}"
        )

    def test_debts_column_count(self, workbook):
        """Verify Debts sheet has exactly 11 column headers"""
        ws = workbook["Debts"]
        headers = [ws.cell(row=2, column=col).value for col in range(1, 13)]
        headers = [h for h in headers if h is not None]
        assert len(headers) == 11, f"Expected 11 headers, got {len(headers)}: {headers}"

    def test_debts_paid_percent_formula(self, workbook):
        """Verify Paid % formula in column 10 references E{row} and F{row}"""
        ws = workbook["Debts"]
        # Check formula in row 3 (first data row)
        formula = ws.cell(row=3, column=10).value
        assert formula is not None, "Column 10 (J) should have a formula"
        assert isinstance(formula, str) and formula.startswith("="), (
            f"Column 10 should contain a formula, got: {formula}"
        )
        # Formula should reference E3 and F3 (Total Loan and Remaining)
        assert "E3" in formula and "F3" in formula, f"Formula should reference E3 and F3, got: {formula}"

    def test_debts_sheet_tab_color(self, workbook):
        """Verify Debts sheet tab color is RED"""
        ws = workbook["Debts"]
        actual_color = ws.sheet_properties.tabColor
        assert actual_color is not None, "Debts sheet should have a tab color"
        color_str = str(actual_color.rgb) if hasattr(actual_color, "rgb") else str(actual_color)
        assert "C00000" in color_str, f"Debts sheet tab color should be RED (C00000), got {color_str}"


class TestDataValidation:
    """Tests for data validation setup"""

    def test_assets_type_has_data_validation(self, workbook):
        """Verify Type column (C) has data validation"""
        ws = workbook["Assets"]
        # Check that data validation exists for this column
        validation_ranges = [dv for dv in ws.data_validations.dataValidation]
        has_type_validation = any("C" in str(dv.sqref) for dv in validation_ranges)
        assert has_type_validation, "Assets sheet column C (Type) should have data validation"

    def test_debts_type_has_data_validation(self, workbook):
        """Verify Type column (C) in Debts has data validation"""
        ws = workbook["Debts"]
        validation_ranges = [dv for dv in ws.data_validations.dataValidation]
        has_type_validation = any("C" in str(dv.sqref) for dv in validation_ranges)
        assert has_type_validation, "Debts sheet column C (Type) should have data validation"

    def test_debts_bank_has_data_validation(self, workbook):
        """Verify Bank column (D) in Debts has data validation"""
        ws = workbook["Debts"]
        validation_ranges = [dv for dv in ws.data_validations.dataValidation]
        has_bank_validation = any("D" in str(dv.sqref) for dv in validation_ranges)
        assert has_bank_validation, "Debts sheet column D (Bank) should have data validation"
