import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from excel_manager import ExcelManager


FIXTURE_WORKBOOK = Path(__file__).resolve().parents[1] / "Financial_Tracker.xlsx"


class ExcelManagerTestCase(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "Financial_Tracker.xlsx"
        shutil.copy2(FIXTURE_WORKBOOK, self.workbook_path)
        self.manager = ExcelManager(self.workbook_path)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_add_asset_writes_to_assets_sheet(self):
        result = self.manager.add_asset(
            name="BCA Reksadana",
            asset_type="Reksadana",
            current_value=5_500_000,
            purchase_value=5_000_000,
            notes="Initial purchase",
        )

        wb = load_workbook(self.workbook_path)
        try:
            ws = wb["Assets"]
            self.assertEqual(result["row"], 3)
            self.assertEqual(ws["B3"].value, "BCA Reksadana")
            self.assertEqual(ws["C3"].value, "Reksadana")
            # D3 = platform (empty string), E3 = purchase_value, F3 = current_value, H3 = notes
            self.assertEqual(ws["E3"].value, 5_000_000)
            self.assertEqual(ws["F3"].value, 5_500_000)
            self.assertEqual(ws["H3"].value, "Initial purchase")
        finally:
            wb.close()

    def test_set_budget_updates_limit_and_month(self):
        self.manager.set_budget("Food & Groceries", 1_250_000, month="2026-03")

        wb = load_workbook(self.workbook_path)
        try:
            ws = wb["Budget"]
            self.assertEqual(ws["B3"].value, 1_250_000)
            self.assertEqual(ws["H2"].value, "2026-03")
        finally:
            wb.close()

    def test_withdrawal_cannot_exceed_current_balance(self):
        with self.assertRaisesRegex(ValueError, "Withdrawal exceeds current balance"):
            self.manager.add_savings(
                amount=2_000_000,
                account="Emergency Fund",
                transaction_type="Withdrawal",
            )


if __name__ == "__main__":
    unittest.main()
