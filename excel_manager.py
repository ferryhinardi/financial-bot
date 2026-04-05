"""
Excel Manager - Read/Write integration with Financial_Tracker.xlsx
"""

from __future__ import annotations

import json
import os
import shutil
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl

from config import local_now


class ExcelManager:
    """Manages all read/write operations to the Financial Tracker Excel file."""

    SHEETS = {
        "transactions": "Transactions",
        "income": "Income",
        "savings": "Savings",
        "assets": "Assets",
        "debts": "Debts",
        "budget": "Budget",
    }

    ROW_RANGES = {
        "transactions": (3, 1002),
        "income": (3, 502),
        "savings": (3, 502),
        "assets": (3, 202),
        "debts": (3, 102),
        "budget": (3, 10),
    }

    BUDGET_MONTH_CELL = "H2"

    CATEGORIES = [
        "Food & Groceries",
        "Transportation",
        "Housing",
        "Entertainment",
        "Healthcare",
        "Education",
        "Shopping",
        "Bills & Utilities",
        "Salary/Income",
    ]

    SAVINGS_ACCOUNTS = [
        "Emergency Fund",
        "Vacation",
        "Investment",
        "Retirement",
        "Other",
    ]

    INCOME_CATEGORIES = [
        "Salary",
        "Freelance",
        "Investment",
        "Side Business",
        "Family Support",
        "Gift",
        "Other",
    ]

    PAYMENT_METHODS = [
        "Cash",
        "Debit Card",
        "Credit Card",
        "Bank Transfer",
        "E-Wallet",
        "Other",
    ]

    ASSET_TYPES = [
        "Ekuitas",
        "Reksadana",
        "Obligasi",
        "Kas",
        "Crypto",
        "Gold",
        "Property",
        "Vehicle",
        "Other",
    ]

    INVESTMENT_PLATFORMS = [
        "Ajaib",
        "Stockbit",
        "Bibit",
        "Bareksa",
        "Manulife",
        "Bank Jago",
        "BSIM3",
        "PRMT2",
        "Other",
    ]

    DEBT_TYPES = [
        "KPR",
        "KTA",
        "Kartu Kredit",
        "KKB",
        "Pinjaman Keluarga",
        "Pinjol",
        "Other",
    ]

    DEBT_BANKS = [
        "BCA",
        "BNI",
        "BRI",
        "Mandiri",
        "BTN",
        "CIMB Niaga",
        "Bank Jago",
        "Other",
    ]

    def __init__(self, excel_path: str | os.PathLike[str], milestones_path: Optional[str | os.PathLike[str]] = None):
        self.excel_path = Path(excel_path).expanduser().resolve()
        if milestones_path is None:
            milestones_path = self.excel_path.parent / "savings_milestones.json"
        self.milestones_path = Path(milestones_path)
        self._ensure_file_exists()

    def _ensure_file_exists(self) -> None:
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

    def _backup(self) -> None:
        """Create a backup before writing and keep only the latest 10 copies."""
        backup_dir = self.excel_path.parent / "backups"
        backup_dir.mkdir(exist_ok=True)

        timestamp = self._current_time().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"Financial_Tracker_backup_{timestamp}.xlsx"
        shutil.copy2(self.excel_path, backup_path)

        backups = sorted(backup_dir.glob("*.xlsx"))
        while len(backups) > 10:
            backups.pop(0).unlink(missing_ok=True)

    def _load_workbook(self):
        return openpyxl.load_workbook(self.excel_path)

    def _save_workbook(self, wb) -> None:
        """Save via a temp file and atomic replace to reduce corruption risk."""
        self._backup()

        temp_fd, temp_name = tempfile.mkstemp(suffix=".xlsx", dir=self.excel_path.parent, prefix=".financial_tracker_")
        os.close(temp_fd)
        temp_path = Path(temp_name)

        try:
            wb.save(temp_path)
            os.replace(temp_path, self.excel_path)
        finally:
            if temp_path.exists():
                temp_path.unlink(missing_ok=True)

    def _find_next_empty_row(self, ws, table_name: str, check_col: int = 1) -> int:
        """Find the next empty row in a worksheet and enforce worksheet limits."""
        start_row, max_row = self.ROW_RANGES[table_name]
        row = start_row
        while row <= max_row and ws.cell(row=row, column=check_col).value is not None:
            row += 1
        if row > max_row:
            raise ValueError(f"{ws.title} sheet is full. Extend the workbook template.")
        return row

    def _current_time(self) -> datetime:
        now = local_now()
        return now.replace(tzinfo=None) if now.tzinfo else now

    def _validate_amount(self, amount: float, *, allow_zero: bool = False) -> float:
        try:
            value = float(amount)
        except (TypeError, ValueError) as exc:
            raise ValueError("Amount must be a number.") from exc

        if allow_zero:
            if value < 0:
                raise ValueError("Amount must not be negative.")
        elif value <= 0:
            raise ValueError("Amount must be greater than zero.")

        return value

    def _validate_choice(self, value: str, allowed: list[str], label: str) -> str:
        normalized = (value or "").strip()
        if normalized not in allowed:
            raise ValueError(f"Invalid {label}: {normalized or '(empty)'}. Allowed values: {', '.join(allowed)}")
        return normalized

    def validate_month(self, month: str) -> str:
        normalized = (month or "").strip()
        if len(normalized) != 7 or normalized[4] != "-":
            raise ValueError("Month must use YYYY-MM format.")
        year, month_part = normalized.split("-", 1)
        if not (year.isdigit() and month_part.isdigit()):
            raise ValueError("Month must use YYYY-MM format.")
        if not 1 <= int(month_part) <= 12:
            raise ValueError("Month must use a valid calendar month.")
        return normalized

    def set_budget_month(self, month: str) -> str:
        """Set the Budget sheet's tracked month (H2 cell). Expects YYYY-MM format."""
        month = self.validate_month(month)
        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["budget"]]
            ws[self.BUDGET_MONTH_CELL] = month
            self._save_workbook(wb)
        finally:
            wb.close()
        return month

    def add_transaction(
        self,
        amount: float,
        category: str,
        description: str = "",
        payment_method: str = "Cash",
        notes: str = "",
        date: Optional[datetime] = None,
    ) -> dict:
        """Add a spending transaction."""
        amount = self._validate_amount(amount)
        category = self._validate_choice(category, self.CATEGORIES, "category")
        payment_method = self._validate_choice(payment_method, self.PAYMENT_METHODS, "payment method")
        date = date or self._current_time()

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["transactions"]]
            row = self._find_next_empty_row(ws, "transactions")

            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
            ws.cell(row=row, column=2, value=(description or "").strip())
            ws.cell(row=row, column=3, value=category)
            ws.cell(row=row, column=4, value=payment_method)
            ws.cell(row=row, column=5, value=amount)
            ws.cell(row=row, column=5).number_format = "#,##0"
            ws.cell(row=row, column=6, value=(notes or "").strip())
            ws.cell(row=row, column=7).value = f'=IF(A{row}<>"",TEXT(A{row},"YYYY-MM"),"")'

            self._save_workbook(wb)
        finally:
            wb.close()

        return {
            "row": row,
            "date": date.strftime("%Y-%m-%d"),
            "amount": amount,
            "category": category,
            "description": (description or "").strip(),
            "payment_method": payment_method,
        }

    def get_spending_summary(self, month: Optional[str] = None) -> dict:
        """Get spending summary, optionally filtered by month (YYYY-MM)."""
        if month is None:
            month = self._current_time().strftime("%Y-%m")
        else:
            month = self.validate_month(month)

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["transactions"]]

            totals = {}
            grand_total = 0.0
            count = 0
            start_row, max_row = self.ROW_RANGES["transactions"]

            for row in range(start_row, max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    break

                if isinstance(date_val, datetime):
                    row_month = date_val.strftime("%Y-%m")
                else:
                    continue

                if row_month != month:
                    continue

                category = ws.cell(row=row, column=3).value or "Uncategorized"
                amount = float(ws.cell(row=row, column=5).value or 0)

                totals[category] = totals.get(category, 0) + amount
                grand_total += amount
                count += 1
        finally:
            wb.close()

        return {
            "month": month,
            "by_category": totals,
            "total": grand_total,
            "transaction_count": count,
        }

    def get_recent_transactions(self, limit: int = 5) -> list:
        """Get the most recent transactions."""
        if limit <= 0:
            raise ValueError("Limit must be greater than zero.")

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["transactions"]]
            transactions = []
            start_row, max_row = self.ROW_RANGES["transactions"]

            for row in range(start_row, max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    break
                transactions.append(
                    {
                        "date": date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val),
                        "description": ws.cell(row=row, column=2).value or "",
                        "category": ws.cell(row=row, column=3).value or "",
                        "amount": float(ws.cell(row=row, column=5).value or 0),
                    }
                )
        finally:
            wb.close()

        return transactions[-limit:]

    # ── Duplicate detection helpers ──

    def find_similar_transactions(
        self,
        amount: float,
        description: str = "",
        date: Optional[datetime] = None,
    ) -> list[dict]:
        """Find existing transactions that look like potential duplicates.

        Matches within the same month where:
        - Amount is within ±1% tolerance, AND
        - Description shares a common substring (case-insensitive, ≥3 chars)

        Returns a list of matching transaction dicts (empty list = no duplicates).
        """
        date = date or self._current_time()
        target_month = date.strftime("%Y-%m")
        desc_lower = (description or "").strip().lower()
        tolerance = abs(amount) * 0.01

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["transactions"]]
            start_row, max_row = self.ROW_RANGES["transactions"]
            matches = []

            for row in range(start_row, max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    break

                if not isinstance(date_val, datetime):
                    continue

                if date_val.strftime("%Y-%m") != target_month:
                    continue

                row_amount = float(ws.cell(row=row, column=5).value or 0)
                if abs(row_amount - amount) > tolerance:
                    continue

                row_desc = (ws.cell(row=row, column=2).value or "").strip().lower()
                descriptions_match = (
                    desc_lower and row_desc and self._has_common_substring(desc_lower, row_desc, min_len=3)
                )
                either_empty = not desc_lower or not row_desc

                if descriptions_match or either_empty:
                    matches.append(
                        {
                            "date": date_val.strftime("%Y-%m-%d"),
                            "description": ws.cell(row=row, column=2).value or "",
                            "category": ws.cell(row=row, column=3).value or "",
                            "amount": row_amount,
                            "payment_method": ws.cell(row=row, column=4).value or "",
                        }
                    )
        finally:
            wb.close()

        return matches

    def find_similar_income(
        self,
        amount: float,
        source: str = "",
        date: Optional[datetime] = None,
    ) -> list[dict]:
        """Find existing income entries that look like potential duplicates.

        Matches within the same month where:
        - Amount is within ±1% tolerance, AND
        - Source shares a common substring (case-insensitive, ≥3 chars)

        Returns a list of matching income dicts (empty list = no duplicates).
        """
        date = date or self._current_time()
        target_month = date.strftime("%Y-%m")
        source_lower = (source or "").strip().lower()
        tolerance = abs(amount) * 0.01

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["income"]]
            start_row, max_row = self.ROW_RANGES["income"]
            matches = []

            for row in range(start_row, max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    break

                if not isinstance(date_val, datetime):
                    continue

                if date_val.strftime("%Y-%m") != target_month:
                    continue

                row_amount = float(ws.cell(row=row, column=4).value or 0)
                if abs(row_amount - amount) > tolerance:
                    continue

                row_source = (ws.cell(row=row, column=2).value or "").strip().lower()
                sources_match = (
                    source_lower and row_source and self._has_common_substring(source_lower, row_source, min_len=3)
                )
                either_empty = not source_lower or not row_source

                if sources_match or either_empty:
                    matches.append(
                        {
                            "date": date_val.strftime("%Y-%m-%d"),
                            "source": ws.cell(row=row, column=2).value or "",
                            "category": ws.cell(row=row, column=3).value or "",
                            "amount": row_amount,
                        }
                    )
        finally:
            wb.close()

        return matches

    @staticmethod
    def _has_common_substring(a: str, b: str, min_len: int = 3) -> bool:
        """Check if two strings share a common substring of at least *min_len* chars.

        Uses word-level matching first (fast path), then falls back to sliding
        window only when necessary.
        """
        words_a = {w for w in a.split() if len(w) >= min_len}
        words_b = {w for w in b.split() if len(w) >= min_len}
        if words_a & words_b:
            return True
        for w in words_a:
            if w in b:
                return True
        for w in words_b:
            if w in a:
                return True
        return False

    # ── Column maps for search / edit / delete ──

    COLUMN_MAP: dict[str, dict[str, int]] = {
        "transactions": {
            "date": 1,
            "description": 2,
            "category": 3,
            "payment_method": 4,
            "amount": 5,
            "notes": 6,
        },
        "income": {
            "date": 1,
            "source": 2,
            "category": 3,
            "amount": 4,
            "notes": 5,
        },
        "savings": {
            "date": 1,
            "account": 2,
            "type": 3,
            "amount": 4,
            "balance": 5,
            "goal": 6,
        },
        "assets": {
            "date": 1,
            "name": 2,
            "type": 3,
            "platform": 4,
            "purchase_value": 5,
            "current_value": 6,
            "notes": 8,
        },
        "debts": {
            "date": 1,
            "name": 2,
            "type": 3,
            "bank": 4,
            "total_loan": 5,
            "remaining": 6,
            "monthly_payment": 7,
            "interest_rate": 8,
            "tenor": 9,
            "notes": 11,
        },
        "budget": {
            "category": 1,
            "budget_limit": 2,
        },
    }

    # Columns that contain formulas — never overwrite via update_row.
    _FORMULA_COLUMNS: dict[str, set[int]] = {
        "transactions": {7},  # Month formula
        "income": {6},  # Month formula
        "savings": {7},  # Progress formula
        "assets": {7},  # Gain/Loss formula
        "debts": {10},  # Paid % formula
        "budget": {3, 4, 5},  # Actual Spent, Remaining, Status formulas
    }

    def _read_row(self, ws, row: int, sheet_key: str) -> dict:
        """Read a single data row into a dict keyed by COLUMN_MAP field names."""
        col_map = self.COLUMN_MAP[sheet_key]
        result: dict = {"_row": row}
        for field, col in col_map.items():
            val = ws.cell(row=row, column=col).value
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            elif val is None:
                val = ""
            result[field] = val
        return result

    @staticmethod
    def _matches_filters(row_data: dict, filters: dict) -> bool:
        """Check whether *row_data* satisfies every entry in *filters*.

        Matching rules:
        - String fields: case-insensitive substring match
        - Numeric fields (amount, balance, etc.): ±1 % tolerance
        - Date fields: exact prefix match (supports "2026", "2026-03", "2026-03-25")
        """
        for key, expected in filters.items():
            actual = row_data.get(key)
            if actual is None or actual == "":
                return False

            try:
                exp_num = float(expected)
                act_num = float(actual)
                tolerance = abs(exp_num) * 0.01
                if abs(act_num - exp_num) > tolerance:
                    return False
                continue
            except (TypeError, ValueError):
                pass

            if key == "date":
                if not str(actual).startswith(str(expected)):
                    return False
                continue

            if str(expected).lower() not in str(actual).lower():
                return False

        return True

    # ── Search / Edit / Delete ──

    def search_rows(self, sheet_key: str, filters: dict, limit: int = 20) -> list[dict]:
        """Return up to *limit* rows in *sheet_key* matching all *filters*."""
        if sheet_key not in self.COLUMN_MAP:
            raise ValueError(f"Unknown sheet key: {sheet_key}. Allowed: {', '.join(self.COLUMN_MAP)}")

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS[sheet_key]]
            start_row, max_row = self.ROW_RANGES[sheet_key]
            matches: list[dict] = []

            for row in range(start_row, max_row + 1):
                if ws.cell(row=row, column=1).value is None:
                    break
                row_data = self._read_row(ws, row, sheet_key)
                if self._matches_filters(row_data, filters):
                    matches.append(row_data)
                    if len(matches) >= limit:
                        break
        finally:
            wb.close()

        return matches

    def update_row(self, sheet_key: str, row_number: int, updates: dict) -> dict:
        """Update fields on a single row; skips formula columns. Returns updated row."""
        if sheet_key not in self.COLUMN_MAP:
            raise ValueError(f"Unknown sheet key: {sheet_key}")
        col_map = self.COLUMN_MAP[sheet_key]
        formula_cols = self._FORMULA_COLUMNS.get(sheet_key, set())
        start_row, max_row = self.ROW_RANGES[sheet_key]

        if not (start_row <= row_number <= max_row):
            raise ValueError(f"Row {row_number} is outside the valid range ({start_row}–{max_row}) for {sheet_key}.")

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS[sheet_key]]

            if ws.cell(row=row_number, column=1).value is None:
                raise ValueError(f"Row {row_number} is empty — nothing to update.")

            for field, new_value in updates.items():
                col = col_map.get(field)
                if col is None:
                    continue
                if col in formula_cols:
                    continue

                if field == "date":
                    if isinstance(new_value, str):
                        new_value = datetime.strptime(new_value, "%Y-%m-%d")
                    ws.cell(row=row_number, column=col, value=new_value)
                    ws.cell(row=row_number, column=col).number_format = "YYYY-MM-DD"
                elif field in (
                    "amount",
                    "balance",
                    "goal",
                    "purchase_value",
                    "current_value",
                    "total_loan",
                    "remaining",
                    "monthly_payment",
                    "interest_rate",
                ):
                    ws.cell(row=row_number, column=col, value=float(new_value))
                    ws.cell(row=row_number, column=col).number_format = "#,##0"
                else:
                    ws.cell(row=row_number, column=col, value=str(new_value))

            self._save_workbook(wb)

            updated = self._read_row(ws, row_number, sheet_key)
        finally:
            wb.close()

        return updated

    def delete_rows(self, sheet_key: str, row_numbers: list[int]) -> int:
        """Delete rows in descending order to preserve numbering. Returns count deleted."""
        if sheet_key not in self.COLUMN_MAP:
            raise ValueError(f"Unknown sheet key: {sheet_key}")
        start_row, max_row = self.ROW_RANGES[sheet_key]

        valid = sorted({r for r in row_numbers if start_row <= r <= max_row}, reverse=True)
        if not valid:
            raise ValueError("No valid row numbers provided.")

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS[sheet_key]]
            deleted = 0
            for row in valid:
                if ws.cell(row=row, column=1).value is None:
                    continue
                ws.delete_rows(row, 1)
                deleted += 1

            if deleted:
                self._save_workbook(wb)
        finally:
            wb.close()

        return deleted

    def update_stock_prices(self) -> list[dict]:
        """Fetch current IDX stock prices via yfinance and update Ekuitas assets.

        Estimates share count from purchase_value / current_price, rounded to
        nearest IDX lot (100 shares), then sets current_value = shares * price.
        """
        import re

        try:
            import yfinance as yf
        except ImportError:
            raise RuntimeError("yfinance is not installed. Run: pip install yfinance")

        equity_rows = self.search_rows("assets", {"type": "Ekuitas"}, limit=50)
        if not equity_rows:
            return []

        ticker_map: dict[str, list[dict]] = {}
        for row in equity_rows:
            name = str(row.get("name", ""))
            match = re.match(r"^([A-Z]{4})\b", name)
            if match:
                ticker = f"{match.group(1)}.JK"
                ticker_map.setdefault(ticker, []).append(row)

        if not ticker_map:
            return []

        results: list[dict] = []
        for ticker, rows in ticker_map.items():
            try:
                stock = yf.Ticker(ticker)
                price_per_share = stock.fast_info.get("lastPrice", 0)
                if not price_per_share or price_per_share <= 0:
                    for row in rows:
                        results.append(
                            {
                                "name": row["name"],
                                "ticker": ticker,
                                "old_value": row.get("current_value", 0),
                                "new_value": 0,
                                "status": "no_price",
                            }
                        )
                    continue

                for row in rows:
                    old_value = float(row.get("current_value", 0) or 0)
                    purchase_value = float(row.get("purchase_value", 0) or 0)

                    if price_per_share > 0 and purchase_value > 0:
                        shares = max(round(purchase_value / price_per_share / 100) * 100, 100)
                        new_value = shares * price_per_share
                    else:
                        shares = 0
                        new_value = old_value

                    try:
                        self.update_row("assets", row["_row"], {"current_value": new_value})
                        results.append(
                            {
                                "name": row["name"],
                                "ticker": ticker,
                                "old_value": old_value,
                                "new_value": new_value,
                                "shares_est": shares,
                                "price_per_share": price_per_share,
                                "status": "updated",
                            }
                        )
                    except Exception as e:
                        results.append(
                            {
                                "name": row["name"],
                                "ticker": ticker,
                                "old_value": old_value,
                                "new_value": new_value,
                                "status": f"update_failed: {e}",
                            }
                        )

            except Exception as e:
                for row in rows:
                    results.append(
                        {
                            "name": row["name"],
                            "ticker": ticker,
                            "old_value": row.get("current_value", 0),
                            "new_value": 0,
                            "status": f"fetch_failed: {e}",
                        }
                    )

        return results

    def add_income(
        self,
        amount: float,
        source: str,
        category: str = "Salary",
        notes: str = "",
        date: Optional[datetime] = None,
    ) -> dict:
        """Add an income entry."""
        amount = self._validate_amount(amount)
        category = self._validate_choice(category, self.INCOME_CATEGORIES, "income category")
        source = (source or "").strip()
        if not source:
            raise ValueError("Income source must not be empty.")
        date = date or self._current_time()

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["income"]]
            row = self._find_next_empty_row(ws, "income")

            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
            ws.cell(row=row, column=2, value=source)
            ws.cell(row=row, column=3, value=category)
            ws.cell(row=row, column=4, value=amount)
            ws.cell(row=row, column=4).number_format = "#,##0"
            ws.cell(row=row, column=5, value=(notes or "").strip())
            ws.cell(row=row, column=6).value = f'=IF(A{row}<>"",TEXT(A{row},"YYYY-MM"),"")'

            self._save_workbook(wb)
        finally:
            wb.close()

        return {
            "row": row,
            "date": date.strftime("%Y-%m-%d"),
            "amount": amount,
            "source": source,
            "category": category,
        }

    def get_income_summary(self, month: Optional[str] = None) -> dict:
        """Get income summary for a month."""
        if month is None:
            month = self._current_time().strftime("%Y-%m")
        else:
            month = self.validate_month(month)

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["income"]]

            totals = {}
            grand_total = 0.0
            start_row, max_row = self.ROW_RANGES["income"]

            for row in range(start_row, max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    break

                if isinstance(date_val, datetime):
                    row_month = date_val.strftime("%Y-%m")
                else:
                    continue

                if row_month != month:
                    continue

                category = ws.cell(row=row, column=3).value or "Other"
                amount = float(ws.cell(row=row, column=4).value or 0)

                totals[category] = totals.get(category, 0) + amount
                grand_total += amount
        finally:
            wb.close()

        return {"month": month, "by_category": totals, "total": grand_total}

    def add_savings(
        self,
        amount: float,
        account: str,
        transaction_type: str = "Deposit",
        goal: Optional[float] = None,
        date: Optional[datetime] = None,
    ) -> dict:
        """Add a savings entry (Deposit, Withdrawal, or Interest)."""
        amount = self._validate_amount(amount)
        account = self._validate_choice(account, self.SAVINGS_ACCOUNTS, "savings account")
        transaction_type = self._validate_choice(
            transaction_type,
            ["Deposit", "Withdrawal", "Interest"],
            "savings transaction type",
        )
        if goal is not None:
            goal = self._validate_amount(goal, allow_zero=True)
        date = date or self._current_time()

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["savings"]]

            current_balance = self._get_account_balance(ws, account)
            if transaction_type == "Withdrawal":
                if amount > current_balance:
                    raise ValueError(f"Withdrawal exceeds current balance for {account}: {current_balance:,.0f}")
                new_balance = current_balance - amount
            else:
                new_balance = current_balance + amount

            row = self._find_next_empty_row(ws, "savings")

            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
            ws.cell(row=row, column=2, value=account)
            ws.cell(row=row, column=3, value=transaction_type)
            ws.cell(row=row, column=4, value=amount)
            ws.cell(row=row, column=4).number_format = "#,##0"
            ws.cell(row=row, column=5, value=new_balance)
            ws.cell(row=row, column=5).number_format = "#,##0"
            if goal is not None:
                ws.cell(row=row, column=6, value=goal)
                ws.cell(row=row, column=6).number_format = "#,##0"
            ws.cell(row=row, column=7).value = f'=IF(AND(E{row}<>"",F{row}<>"",F{row}<>0),E{row}/F{row},"")'
            ws.cell(row=row, column=7).number_format = "0.0%"

            self._save_workbook(wb)
        finally:
            wb.close()

        return {
            "row": row,
            "date": date.strftime("%Y-%m-%d"),
            "account": account,
            "type": transaction_type,
            "amount": amount,
            "balance": new_balance,
        }

    def _get_account_balance(self, ws, account: str) -> float:
        """Get the latest balance for a savings account."""
        balance = 0.0
        start_row, max_row = self.ROW_RANGES["savings"]
        for row in range(start_row, max_row + 1):
            if ws.cell(row=row, column=1).value is None:
                break
            if ws.cell(row=row, column=2).value == account:
                val = ws.cell(row=row, column=5).value
                if val is not None:
                    balance = float(val)
        return balance

    def get_savings_summary(self) -> dict:
        """Get current savings balances for all accounts."""
        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["savings"]]

            accounts = {}
            start_row, max_row = self.ROW_RANGES["savings"]
            for row in range(start_row, max_row + 1):
                if ws.cell(row=row, column=1).value is None:
                    break
                account = ws.cell(row=row, column=2).value
                if account:
                    balance = float(ws.cell(row=row, column=5).value or 0)
                    goal = ws.cell(row=row, column=6).value
                    accounts[account] = {
                        "balance": balance,
                        "goal": float(goal) if goal is not None else None,
                    }
        finally:
            wb.close()

        total = sum(a["balance"] for a in accounts.values())
        return {"accounts": accounts, "total_savings": total}

    def add_asset(
        self,
        name: str,
        asset_type: str,
        current_value: float,
        purchase_value: Optional[float] = None,
        platform: str = "",
        notes: str = "",
        date: Optional[datetime] = None,
    ) -> dict:
        """Add an asset or investment entry."""
        name = (name or "").strip()
        if not name:
            raise ValueError("Asset name must not be empty.")

        asset_type = self._validate_choice(asset_type, self.ASSET_TYPES, "asset type")
        current_value = self._validate_amount(current_value, allow_zero=True)
        if purchase_value is None:
            purchase_value = current_value
        purchase_value = self._validate_amount(purchase_value, allow_zero=True)
        date = date or self._current_time()

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["assets"]]
            row = self._find_next_empty_row(ws, "assets")

            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=asset_type)
            ws.cell(row=row, column=4, value=(platform or "").strip())
            ws.cell(row=row, column=5, value=purchase_value)
            ws.cell(row=row, column=5).number_format = "#,##0"
            ws.cell(row=row, column=6, value=current_value)
            ws.cell(row=row, column=6).number_format = "#,##0"
            ws.cell(row=row, column=7).value = f'=IF(AND(E{row}<>"",F{row}<>""),F{row}-E{row},"")'
            ws.cell(row=row, column=7).number_format = "#,##0"
            ws.cell(row=row, column=8, value=(notes or "").strip())

            self._save_workbook(wb)
        finally:
            wb.close()

        return {
            "row": row,
            "date": date.strftime("%Y-%m-%d"),
            "name": name,
            "type": asset_type,
            "platform": (platform or "").strip(),
            "purchase_value": purchase_value,
            "current_value": current_value,
        }

    def get_investment_summary(self) -> dict:
        """Get a summary of all investment assets grouped by type and platform."""
        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["assets"]]
            start_row, max_row = self.ROW_RANGES["assets"]

            assets = []
            by_type = {}
            by_platform = {}
            grand_total_current = 0.0
            grand_total_purchase = 0.0

            for row in range(start_row, max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    break

                name = ws.cell(row=row, column=2).value or ""
                asset_type = ws.cell(row=row, column=3).value or "Other"
                platform = ws.cell(row=row, column=4).value or ""
                purchase_val = float(ws.cell(row=row, column=5).value or 0)
                current_val = float(ws.cell(row=row, column=6).value or 0)
                notes = ws.cell(row=row, column=8).value or ""

                asset = {
                    "date": date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val),
                    "name": name,
                    "type": asset_type,
                    "platform": platform,
                    "purchase_value": purchase_val,
                    "current_value": current_val,
                    "gain_loss": current_val - purchase_val,
                    "notes": notes,
                }
                assets.append(asset)

                # Group by type
                if asset_type not in by_type:
                    by_type[asset_type] = {
                        "current_value": 0.0,
                        "purchase_value": 0.0,
                        "items": [],
                    }
                by_type[asset_type]["current_value"] += current_val
                by_type[asset_type]["purchase_value"] += purchase_val
                by_type[asset_type]["items"].append(asset)

                # Group by platform
                if platform:
                    if platform not in by_platform:
                        by_platform[platform] = {"current_value": 0.0, "items": []}
                    by_platform[platform]["current_value"] += current_val
                    by_platform[platform]["items"].append(asset)

                grand_total_current += current_val
                grand_total_purchase += purchase_val
        finally:
            wb.close()

        return {
            "assets": assets,
            "by_type": by_type,
            "by_platform": by_platform,
            "total_current": grand_total_current,
            "total_purchase": grand_total_purchase,
            "total_gain_loss": grand_total_current - grand_total_purchase,
            "asset_count": len(assets),
        }

    def set_budget(self, category: str, amount: float, month: Optional[str] = None) -> dict:
        """Update the budget limit for a category and optionally the tracked month."""
        category = self._validate_choice(category, self.CATEGORIES, "budget category")
        amount = self._validate_amount(amount, allow_zero=True)
        if month is not None:
            month = self.validate_month(month)

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["budget"]]
            row = self._get_budget_row(ws, category)
            ws.cell(row=row, column=2, value=amount)
            ws.cell(row=row, column=2).number_format = "#,##0"
            if month is not None:
                ws[self.BUDGET_MONTH_CELL] = month

            self._save_workbook(wb)
        finally:
            wb.close()

        return {"category": category, "amount": amount, "month": month}

    def _get_budget_row(self, ws, category: str) -> int:
        start_row, max_row = self.ROW_RANGES["budget"]
        for row in range(start_row, max_row + 1):
            if ws.cell(row=row, column=1).value == category:
                return row
        raise ValueError(f"Category {category} not found on Budget sheet.")

    def get_budget_status(self, month: Optional[str] = None) -> dict:
        """Get budget status for a month."""
        if month is None:
            month = self._current_time().strftime("%Y-%m")
        else:
            month = self.validate_month(month)

        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["budget"]]
            spending = self.get_spending_summary(month)

            budget_items = []
            start_row, max_row = self.ROW_RANGES["budget"]
            for row in range(start_row, max_row + 1):
                category = ws.cell(row=row, column=1).value
                if not category:
                    continue

                budget_limit = float(ws.cell(row=row, column=2).value or 0)
                actual = float(spending["by_category"].get(category, 0))
                remaining = budget_limit - actual
                pct = (actual / budget_limit * 100) if budget_limit > 0 else 0

                if pct > 100:
                    status = "OVER BUDGET"
                elif pct > 80:
                    status = "WARNING"
                else:
                    status = "OK"

                budget_items.append(
                    {
                        "category": category,
                        "budget": budget_limit,
                        "spent": actual,
                        "remaining": remaining,
                        "percentage": pct,
                        "status": status,
                    }
                )
        finally:
            wb.close()

        total_budget = sum(b["budget"] for b in budget_items)
        total_spent = sum(b["spent"] for b in budget_items)

        return {
            "month": month,
            "items": budget_items,
            "total_budget": total_budget,
            "total_spent": total_spent,
            "total_remaining": total_budget - total_spent,
        }

    def get_dashboard(self, month: Optional[str] = None) -> dict:
        """Get a full financial dashboard summary."""
        if month is None:
            month = self._current_time().strftime("%Y-%m")
        else:
            month = self.validate_month(month)

        income = self.get_income_summary(month)
        spending = self.get_spending_summary(month)
        savings = self.get_savings_summary()
        budget = self.get_budget_status(month)
        investments = self.get_investment_summary()
        debts = self.get_debt_summary()

        return {
            "month": month,
            "income": income["total"],
            "spending": spending["total"],
            "net": income["total"] - spending["total"],
            "savings_total": savings["total_savings"],
            "budget_remaining": budget["total_remaining"],
            "spending_by_category": spending["by_category"],
            "savings_accounts": savings["accounts"],
            "investment_total": investments["total_current"],
            "investment_gain_loss": investments["total_gain_loss"],
            "debt_total": debts["total_remaining"],
            "net_worth": savings["total_savings"] + investments["total_current"] - debts["total_remaining"],
        }

    # ── Debt / Liability tracking ──

    def add_debt(
        self,
        name: str,
        debt_type: str,
        bank: str,
        total_loan: float,
        remaining: float,
        monthly_payment: float,
        interest_rate: float = 0.0,
        tenor_months: int = 0,
        notes: str = "",
        date: Optional[datetime] = None,
    ) -> dict:
        """Add a debt or liability entry (KPR, KTA, credit card, etc.)."""
        name = (name or "").strip()
        if not name:
            raise ValueError("Debt name must not be empty.")

        debt_type = self._validate_choice(debt_type, self.DEBT_TYPES, "debt type")
        total_loan = self._validate_amount(total_loan)
        remaining = self._validate_amount(remaining, allow_zero=True)
        monthly_payment = self._validate_amount(monthly_payment, allow_zero=True)
        date = date or self._current_time()

        wb = self._load_workbook()
        try:
            # Create Debts sheet if it doesn't exist
            if self.SHEETS["debts"] not in wb.sheetnames:
                ws = wb.create_sheet(self.SHEETS["debts"])
                headers = [
                    "Date",
                    "Name",
                    "Type",
                    "Bank",
                    "Total Loan",
                    "Remaining",
                    "Monthly Payment",
                    "Interest Rate",
                    "Tenor (months)",
                    "Paid %",
                    "Notes",
                ]
                for col, h in enumerate(headers, 1):
                    ws.cell(row=2, column=col, value=h)
            else:
                ws = wb[self.SHEETS["debts"]]

            row = self._find_next_empty_row(ws, "debts")

            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=debt_type)
            ws.cell(row=row, column=4, value=(bank or "").strip())
            ws.cell(row=row, column=5, value=total_loan)
            ws.cell(row=row, column=5).number_format = "#,##0"
            ws.cell(row=row, column=6, value=remaining)
            ws.cell(row=row, column=6).number_format = "#,##0"
            ws.cell(row=row, column=7, value=monthly_payment)
            ws.cell(row=row, column=7).number_format = "#,##0"
            ws.cell(row=row, column=8, value=interest_rate)
            ws.cell(row=row, column=8).number_format = "0.00%"
            ws.cell(row=row, column=9, value=tenor_months)
            ws.cell(row=row, column=10).value = f'=IF(AND(E{row}<>"",F{row}<>"",E{row}<>0),(E{row}-F{row})/E{row},"")'
            ws.cell(row=row, column=10).number_format = "0.0%"
            ws.cell(row=row, column=11, value=(notes or "").strip())

            self._save_workbook(wb)
        finally:
            wb.close()

        paid_pct = ((total_loan - remaining) / total_loan * 100) if total_loan > 0 else 0

        return {
            "row": row,
            "date": date.strftime("%Y-%m-%d"),
            "name": name,
            "type": debt_type,
            "bank": (bank or "").strip(),
            "total_loan": total_loan,
            "remaining": remaining,
            "monthly_payment": monthly_payment,
            "interest_rate": interest_rate,
            "tenor_months": tenor_months,
            "paid_pct": paid_pct,
        }

    def get_debt_summary(self) -> dict:
        """Get a summary of all debts/liabilities."""
        wb = self._load_workbook()
        try:
            if self.SHEETS["debts"] not in wb.sheetnames:
                return {
                    "debts": [],
                    "by_type": {},
                    "total_loan": 0.0,
                    "total_remaining": 0.0,
                    "total_monthly": 0.0,
                    "debt_count": 0,
                }

            ws = wb[self.SHEETS["debts"]]
            start_row, max_row = self.ROW_RANGES["debts"]

            debts = []
            by_type = {}
            grand_total_loan = 0.0
            grand_total_remaining = 0.0
            grand_total_monthly = 0.0

            for row in range(start_row, max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    break

                name = ws.cell(row=row, column=2).value or ""
                debt_type = ws.cell(row=row, column=3).value or "Other"
                bank = ws.cell(row=row, column=4).value or ""
                total_loan = float(ws.cell(row=row, column=5).value or 0)
                remaining = float(ws.cell(row=row, column=6).value or 0)
                monthly_payment = float(ws.cell(row=row, column=7).value or 0)
                interest_rate = float(ws.cell(row=row, column=8).value or 0)
                tenor_months = int(ws.cell(row=row, column=9).value or 0)
                notes = ws.cell(row=row, column=11).value or ""

                paid_pct = ((total_loan - remaining) / total_loan * 100) if total_loan > 0 else 0

                debt = {
                    "date": date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val),
                    "name": name,
                    "type": debt_type,
                    "bank": bank,
                    "total_loan": total_loan,
                    "remaining": remaining,
                    "monthly_payment": monthly_payment,
                    "interest_rate": interest_rate,
                    "tenor_months": tenor_months,
                    "paid_pct": paid_pct,
                    "notes": notes,
                }
                debts.append(debt)

                if debt_type not in by_type:
                    by_type[debt_type] = {
                        "total_loan": 0.0,
                        "remaining": 0.0,
                        "items": [],
                    }
                by_type[debt_type]["total_loan"] += total_loan
                by_type[debt_type]["remaining"] += remaining
                by_type[debt_type]["items"].append(debt)

                grand_total_loan += total_loan
                grand_total_remaining += remaining
                grand_total_monthly += monthly_payment
        finally:
            wb.close()

        return {
            "debts": debts,
            "by_type": by_type,
            "total_loan": grand_total_loan,
            "total_remaining": grand_total_remaining,
            "total_monthly": grand_total_monthly,
            "debt_count": len(debts),
        }

    def get_savings_goals(self) -> list[dict]:
        wb = self._load_workbook()
        try:
            ws = wb[self.SHEETS["savings"]]

            account_data = {}
            start_row, max_row = self.ROW_RANGES["savings"]

            for row in range(start_row, max_row + 1):
                if ws.cell(row=row, column=1).value is None:
                    break
                date_val = ws.cell(row=row, column=1).value
                account = ws.cell(row=row, column=2).value
                amount = ws.cell(row=row, column=4).value or 0
                balance = ws.cell(row=row, column=5).value or 0
                goal = ws.cell(row=row, column=6).value

                if account:
                    if account not in account_data:
                        account_data[account] = {"balance": 0, "goal": goal, "deposits": []}

                    account_data[account]["balance"] = float(balance)
                    if isinstance(date_val, datetime):
                        account_data[account]["deposits"].append({"date": date_val, "amount": float(amount)})
        finally:
            wb.close()

        result = []
        now = self._current_time()

        for account, data in account_data.items():
            balance = data["balance"]
            goal = float(data["goal"]) if data["goal"] else 0

            progress_pct = (balance / goal * 100) if goal > 0 else 0

            milestones_hit = []
            for milestone in [25, 50, 75, 90, 100]:
                if progress_pct >= milestone:
                    milestones_hit.append(milestone)

            eta_months = None
            deposits = data["deposits"]
            if deposits and goal > 0 and balance < goal:
                three_months_ago = now - timedelta(days=90)
                recent = [d for d in deposits if d["date"] >= three_months_ago]
                if recent:
                    avg_monthly = sum(d["amount"] for d in recent) / 3
                    if avg_monthly > 0:
                        eta_months = (goal - balance) / avg_monthly

            result.append(
                {
                    "account": account,
                    "balance": balance,
                    "goal": goal,
                    "progress_pct": round(progress_pct, 1),
                    "eta_months": round(eta_months, 1) if eta_months else None,
                    "milestones_hit": milestones_hit,
                }
            )

        return result

    def check_milestone(self, account: str) -> Optional[dict]:
        milestones = self._load_milestones()

        goals = self.get_savings_goals()
        goal_data = next((g for g in goals if g["account"] == account), None)

        if not goal_data:
            return None

        current_pct = goal_data["progress_pct"]
        last_milestone = milestones.get(account, {}).get("last_milestone", 0)

        for milestone in [25, 50, 75, 90, 100]:
            if current_pct >= milestone and milestone > last_milestone:
                milestones[account] = {
                    "last_milestone": milestone,
                    "checked_at": self._current_time().strftime("%Y-%m"),
                }
                self._save_milestones(milestones)

                return {
                    "account": account,
                    "milestone_pct": milestone,
                    "message": self.get_celebration_message(milestone),
                }

        return None

    def get_celebration_message(self, milestone_pct: int) -> str:
        messages = {
            25: "🎉 Hebat! Kamu sudah mencapai 25% dari targetmu!",
            50: "🚀 Luar biasa! Setengah perjalanan sudah tercapai!",
            75: "💪 Hampir sampai! 75% target tercapai!",
            90: "⭐ Sedikit lagi! Tinggal 10% menuju target!",
            100: "🏆 SELAMAT! Target tabungan tercapai!",
        }
        return messages.get(milestone_pct, f"✨ Milestone {milestone_pct}% tercapai!")

    def _load_milestones(self) -> dict:
        if self.milestones_path.exists():
            with open(self.milestones_path) as f:
                return json.load(f)
        return {}

    def _save_milestones(self, data: dict) -> None:
        with open(self.milestones_path, "w") as f:
            json.dump(data, f, indent=2)
