import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

wb = openpyxl.Workbook()

# ── Color Palette ──
DARK_BLUE = "1F4E79"
BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
GREEN = "548235"
LIGHT_GREEN = "E2EFDA"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
RED = "C00000"
LIGHT_RED = "FBE5D6"
GRAY = "D9D9D9"
WHITE = "FFFFFF"
DARK_GRAY = "404040"
PURPLE = "7030A0"
LIGHT_PURPLE = "E2D0F0"
DARK_GREEN = "375623"
YELLOW_BG = "FFF2CC"

# ── Styles ──
header_font = Font(name="Calibri", size=12, bold=True, color=WHITE)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
subheader_font = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
subheader_fill = PatternFill(
    start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"
)
title_font = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
currency_format = "#,##0"
percent_format = "0.0%"
date_format = "YYYY-MM-DD"
thin_border = Border(
    left=Side(style="thin", color=GRAY),
    right=Side(style="thin", color=GRAY),
    top=Side(style="thin", color=GRAY),
    bottom=Side(style="thin", color=GRAY),
)

CATEGORIES = [
    "Food & Groceries",
    "Transportation",
    "Housing",
    "Entertainment",
    "Healthcare",
    "Education",
    "Shopping",
    "Bills & Utilities",
]

SAVINGS_ACCOUNTS = [
    "Emergency Fund",
    "Vacation",
    "Investment",
    "Retirement",
    "Other",
]

ASSET_TYPES = [
    "Stocks",
    "Mutual Fund",
    "Crypto",
    "Gold",
    "Property",
    "Vehicle",
    "Other",
]


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border


def section_title(ws, row, start_col, end_col, text, color=BLUE):
    ws.merge_cells(
        start_row=row, start_column=start_col, end_row=row, end_column=end_col
    )
    cell = ws.cell(row=row, column=start_col)
    cell.value = text
    cell.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = fill


def style_data_rows(ws, start_row, end_row, start_col, end_col, alt_fill=None):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if alt_fill and (r - start_row) % 2 == 0:
                cell.fill = alt_fill


# ═══════════════════════════════════════════════════════════════
# SHEET 1: TRANSACTIONS
# ═══════════════════════════════════════════════════════════════
ws_txn = wb.active
ws_txn.title = "Transactions"
ws_txn.sheet_properties.tabColor = BLUE

ws_txn.merge_cells("A1:G1")
ws_txn["A1"].value = "SPENDING TRACKER"
ws_txn["A1"].font = title_font
ws_txn["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_txn.row_dimensions[1].height = 35

txn_headers = [
    "Date",
    "Description",
    "Category",
    "Payment Method",
    "Amount",
    "Notes",
    "Month",
]
for col, h in enumerate(txn_headers, 1):
    ws_txn.cell(row=2, column=col, value=h)
style_header_row(ws_txn, 2, len(txn_headers))

for i, w in enumerate([14, 30, 20, 18, 15, 25, 12], 1):
    ws_txn.column_dimensions[get_column_letter(i)].width = w

cat_validation = DataValidation(
    type="list", formula1='"' + ",".join(CATEGORIES) + '"', allow_blank=True
)
cat_validation.error = "Please select a valid category"
ws_txn.add_data_validation(cat_validation)
cat_validation.add("C3:C1002")

payment_methods = [
    "Cash",
    "Debit Card",
    "Credit Card",
    "Bank Transfer",
    "E-Wallet",
    "Other",
]
pay_validation = DataValidation(
    type="list", formula1='"' + ",".join(payment_methods) + '"', allow_blank=True
)
ws_txn.add_data_validation(pay_validation)
pay_validation.add("D3:D1002")

for row in range(3, 1003):
    ws_txn.cell(row=row, column=1).number_format = date_format
    ws_txn.cell(row=row, column=5).number_format = currency_format
    ws_txn.cell(row=row, column=7).value = f'=IF(A{row}<>"",TEXT(A{row},"YYYY-MM"),"")'
    ws_txn.cell(row=row, column=7).number_format = "@"
    for c in range(1, 8):
        ws_txn.cell(row=row, column=c).border = thin_border

samples = [
    (
        datetime(2026, 1, 5),
        "Weekly groceries",
        "Food & Groceries",
        "Debit Card",
        85000,
        "Supermarket",
    ),
    (
        datetime(2026, 1, 6),
        "Bus pass monthly",
        "Transportation",
        "E-Wallet",
        50000,
        "Monthly pass",
    ),
    (
        datetime(2026, 1, 7),
        "Netflix subscription",
        "Entertainment",
        "Credit Card",
        54000,
        "Monthly",
    ),
]
for i, (dt, desc, cat, pay, amt, note) in enumerate(samples, 3):
    ws_txn.cell(row=i, column=1, value=dt)
    ws_txn.cell(row=i, column=2, value=desc)
    ws_txn.cell(row=i, column=3, value=cat)
    ws_txn.cell(row=i, column=4, value=pay)
    ws_txn.cell(row=i, column=5, value=amt)
    ws_txn.cell(row=i, column=6, value=note)
ws_txn.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════
# SHEET 2: INCOME
# ═══════════════════════════════════════════════════════════════
ws_income = wb.create_sheet("Income")
ws_income.sheet_properties.tabColor = GREEN

ws_income.merge_cells("A1:F1")
ws_income["A1"].value = "INCOME TRACKER"
ws_income["A1"].font = title_font
ws_income["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_income.row_dimensions[1].height = 35

for col, h in enumerate(["Date", "Source", "Category", "Amount", "Notes", "Month"], 1):
    ws_income.cell(row=2, column=col, value=h)
style_header_row(ws_income, 2, 6)

for i, w in enumerate([14, 25, 18, 15, 30, 12], 1):
    ws_income.column_dimensions[get_column_letter(i)].width = w

income_cats = ["Salary", "Freelance", "Investment", "Side Business", "Gift", "Other"]
inc_val = DataValidation(
    type="list", formula1='"' + ",".join(income_cats) + '"', allow_blank=True
)
ws_income.add_data_validation(inc_val)
inc_val.add("C3:C502")

for row in range(3, 503):
    ws_income.cell(row=row, column=1).number_format = date_format
    ws_income.cell(row=row, column=4).number_format = currency_format
    ws_income.cell(
        row=row, column=6
    ).value = f'=IF(A{row}<>"",TEXT(A{row},"YYYY-MM"),"")'
    for c in range(1, 7):
        ws_income.cell(row=row, column=c).border = thin_border

ws_income.cell(row=3, column=1, value=datetime(2026, 1, 1))
ws_income.cell(row=3, column=2, value="Monthly Salary")
ws_income.cell(row=3, column=3, value="Salary")
ws_income.cell(row=3, column=4, value=5000000)
ws_income.cell(row=3, column=5, value="January salary")
ws_income.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════
# SHEET 3: SAVINGS
# ═══════════════════════════════════════════════════════════════
ws_sav = wb.create_sheet("Savings")
ws_sav.sheet_properties.tabColor = "548235"

ws_sav.merge_cells("A1:G1")
ws_sav["A1"].value = "SAVINGS TRACKER"
ws_sav["A1"].font = title_font
ws_sav["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sav.row_dimensions[1].height = 35

for col, h in enumerate(
    ["Date", "Account", "Type", "Amount", "Balance After", "Goal", "Progress"], 1
):
    ws_sav.cell(row=2, column=col, value=h)
style_header_row(ws_sav, 2, 7)

for i, w in enumerate([14, 20, 15, 15, 18, 18, 14], 1):
    ws_sav.column_dimensions[get_column_letter(i)].width = w

acct_val = DataValidation(
    type="list", formula1='"' + ",".join(SAVINGS_ACCOUNTS) + '"', allow_blank=True
)
ws_sav.add_data_validation(acct_val)
acct_val.add("B3:B502")

type_val = DataValidation(
    type="list", formula1='"Deposit,Withdrawal,Interest"', allow_blank=True
)
ws_sav.add_data_validation(type_val)
type_val.add("C3:C502")

for row in range(3, 503):
    ws_sav.cell(row=row, column=1).number_format = date_format
    ws_sav.cell(row=row, column=4).number_format = currency_format
    ws_sav.cell(row=row, column=5).number_format = currency_format
    ws_sav.cell(row=row, column=6).number_format = currency_format
    ws_sav.cell(row=row, column=7).number_format = percent_format
    ws_sav.cell(
        row=row, column=7
    ).value = f'=IF(AND(E{row}<>"",F{row}<>"",F{row}<>0),E{row}/F{row},"")'
    for c in range(1, 8):
        ws_sav.cell(row=row, column=c).border = thin_border

sav_samples = [
    (datetime(2026, 1, 1), "Emergency Fund", "Deposit", 1000000, 1000000, 10000000),
    (datetime(2026, 1, 1), "Vacation", "Deposit", 500000, 500000, 3000000),
    (datetime(2026, 1, 1), "Investment", "Deposit", 2000000, 2000000, 20000000),
    (datetime(2026, 1, 15), "Emergency Fund", "Deposit", 500000, 1500000, 10000000),
]
for i, (dt, acct, typ, amt, bal, goal) in enumerate(sav_samples, 3):
    ws_sav.cell(row=i, column=1, value=dt)
    ws_sav.cell(row=i, column=2, value=acct)
    ws_sav.cell(row=i, column=3, value=typ)
    ws_sav.cell(row=i, column=4, value=amt)
    ws_sav.cell(row=i, column=5, value=bal)
    ws_sav.cell(row=i, column=6, value=goal)
ws_sav.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════
# SHEET 4: ASSETS
# ═══════════════════════════════════════════════════════════════
ws_assets = wb.create_sheet("Assets")
ws_assets.sheet_properties.tabColor = PURPLE

ws_assets.merge_cells("A1:G1")
ws_assets["A1"].value = "ASSETS & INVESTMENTS"
ws_assets["A1"].font = title_font
ws_assets["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_assets.row_dimensions[1].height = 35

asset_headers = [
    "Date Added",
    "Name",
    "Type",
    "Purchase Value",
    "Current Value",
    "Gain/Loss",
    "Notes",
]
for col, h in enumerate(asset_headers, 1):
    ws_assets.cell(row=2, column=col, value=h)
style_header_row(ws_assets, 2, len(asset_headers))

for i, w in enumerate([14, 28, 18, 18, 18, 16, 25], 1):
    ws_assets.column_dimensions[get_column_letter(i)].width = w

atype_val = DataValidation(
    type="list", formula1='"' + ",".join(ASSET_TYPES) + '"', allow_blank=True
)
ws_assets.add_data_validation(atype_val)
atype_val.add("C3:C202")

for row in range(3, 203):
    ws_assets.cell(row=row, column=1).number_format = date_format
    ws_assets.cell(row=row, column=4).number_format = currency_format
    ws_assets.cell(row=row, column=5).number_format = currency_format
    ws_assets.cell(row=row, column=6).number_format = currency_format
    ws_assets.cell(
        row=row, column=6
    ).value = f'=IF(AND(D{row}<>"",E{row}<>""),E{row}-D{row},"")'
    for c in range(1, 8):
        ws_assets.cell(row=row, column=c).border = thin_border
ws_assets.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════
# SHEET 5: BUDGET
# ═══════════════════════════════════════════════════════════════
ws_budget = wb.create_sheet("Budget")
ws_budget.sheet_properties.tabColor = RED

ws_budget.merge_cells("A1:E1")
ws_budget["A1"].value = "MONTHLY BUDGET PLANNER"
ws_budget["A1"].font = title_font
ws_budget["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_budget.row_dimensions[1].height = 35

for col, h in enumerate(
    ["Category", "Budget Limit", "Actual Spent", "Remaining", "Status"], 1
):
    ws_budget.cell(row=2, column=col, value=h)
style_header_row(ws_budget, 2, 5)

for i, w in enumerate([22, 18, 18, 18, 15], 1):
    ws_budget.column_dimensions[get_column_letter(i)].width = w

ws_budget.cell(row=2, column=7, value="Budget Month:").font = subheader_font
ws_budget.cell(row=2, column=8, value="2026-01").font = Font(
    size=12, bold=True, color=RED
)
ws_budget.column_dimensions["G"].width = 16
ws_budget.column_dimensions["H"].width = 14

default_budgets = [500000, 200000, 1500000, 200000, 300000, 200000, 300000, 500000]
for i, cat in enumerate(CATEGORIES):
    r = 3 + i
    ws_budget.cell(row=r, column=1, value=cat)
    ws_budget.cell(
        row=r, column=2, value=default_budgets[i]
    ).number_format = currency_format
    ws_budget.cell(row=r, column=3).value = (
        f"=SUMPRODUCT((EXACT(Transactions!C3:C1002,A{r}))"
        f"*(EXACT(Transactions!G3:G1002,$H$2))"
        f"*(Transactions!E3:E1002))"
    )
    ws_budget.cell(row=r, column=3).number_format = currency_format
    ws_budget.cell(row=r, column=4).value = f'=IF(B{r}<>"",B{r}-C{r},"")'
    ws_budget.cell(row=r, column=4).number_format = currency_format
    ws_budget.cell(
        row=r, column=5
    ).value = (
        f'=IF(B{r}<>"",IF(C{r}>B{r},"OVER BUDGET",IF(C{r}>B{r}*0.8,"WARNING","OK")),"")'
    )
    for c in range(1, 6):
        ws_budget.cell(row=r, column=c).border = thin_border

total_budget_row = 3 + len(CATEGORIES)
ws_budget.cell(row=total_budget_row, column=1, value="TOTAL").font = Font(bold=True)
ws_budget.cell(
    row=total_budget_row, column=2
).value = f"=SUM(B3:B{total_budget_row - 1})"
ws_budget.cell(row=total_budget_row, column=2).number_format = currency_format
ws_budget.cell(
    row=total_budget_row, column=3
).value = f"=SUM(C3:C{total_budget_row - 1})"
ws_budget.cell(row=total_budget_row, column=3).number_format = currency_format
ws_budget.cell(
    row=total_budget_row, column=4
).value = f"=B{total_budget_row}-C{total_budget_row}"
ws_budget.cell(row=total_budget_row, column=4).number_format = currency_format
for c in range(1, 6):
    ws_budget.cell(row=total_budget_row, column=c).border = thin_border
    ws_budget.cell(row=total_budget_row, column=c).fill = PatternFill(
        start_color=GRAY, end_color=GRAY, fill_type="solid"
    )

# ═══════════════════════════════════════════════════════════════
# SHEET 6: DASHBOARD
# ═══════════════════════════════════════════════════════════════
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_properties.tabColor = ORANGE

ws_dash.merge_cells("A1:I1")
ws_dash["A1"].value = "FINANCIAL DASHBOARD"
ws_dash["A1"].font = Font(name="Calibri", size=18, bold=True, color=DARK_BLUE)
ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[1].height = 40

# ── Financial Overview ──
row = 3
section_title(ws_dash, row, 1, 4, "FINANCIAL OVERVIEW", BLUE)

overview = [
    ("Total Income", "=SUMPRODUCT((Income!D3:D502)*1)"),
    ("Total Spending", "=SUMPRODUCT((Transactions!E3:E1002)*1)"),
    ("Net (Income-Spend)", "=B4-B5"),
    (
        "Total Savings",
        '=SUMPRODUCT((EXACT(Savings!C3:C502,"Deposit"))*(Savings!D3:D502))-SUMPRODUCT((EXACT(Savings!C3:C502,"Withdrawal"))*(Savings!D3:D502))',
    ),
    ("Total Assets", "=SUMPRODUCT((Assets!E3:E202)*1)"),
    ("NET WORTH", "=B7+B8"),
]
for i, (label, formula) in enumerate(overview):
    r = row + 1 + i
    ws_dash.cell(row=r, column=1, value=label).font = subheader_font
    ws_dash.cell(row=r, column=1).fill = subheader_fill
    ws_dash.cell(row=r, column=2, value=formula).number_format = currency_format
    ws_dash.cell(row=r, column=2).font = Font(name="Calibri", size=12, bold=True)
    for c in range(1, 5):
        ws_dash.cell(row=r, column=c).border = thin_border

# Highlight net and net worth
for highlight_row in [6, 9]:
    for c in range(1, 5):
        ws_dash.cell(row=highlight_row, column=c).fill = PatternFill(
            start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid"
        )
ws_dash.cell(row=9, column=1).font = Font(
    name="Calibri", size=12, bold=True, color=DARK_GREEN
)
ws_dash.cell(row=9, column=2).font = Font(
    name="Calibri", size=14, bold=True, color=DARK_GREEN
)

# ── Spending by Category ──
cat_section_row = 11
section_title(ws_dash, cat_section_row, 1, 4, "SPENDING BY CATEGORY", BLUE)

cat_h = cat_section_row + 1
for col, h in enumerate(["Category", "Total Spent", "% of Total", "Avg per Txn"], 1):
    ws_dash.cell(row=cat_h, column=col, value=h)
style_header_row(ws_dash, cat_h, 4)

alt_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
for i, cat in enumerate(CATEGORIES):
    r = cat_h + 1 + i
    ws_dash.cell(row=r, column=1, value=cat).font = Font(name="Calibri", size=11)
    ws_dash.cell(
        row=r, column=2
    ).value = (
        f"=SUMPRODUCT((EXACT(Transactions!C3:C1002,A{r}))*(Transactions!E3:E1002))"
    )
    ws_dash.cell(row=r, column=2).number_format = currency_format
    ws_dash.cell(row=r, column=3).value = f"=IF(B5<>0,B{r}/B5,0)"
    ws_dash.cell(row=r, column=3).number_format = percent_format
    ws_dash.cell(
        row=r, column=4
    ).value = f"=IF(COUNTIF(Transactions!C3:C1002,A{r})>0,B{r}/COUNTIF(Transactions!C3:C1002,A{r}),0)"
    ws_dash.cell(row=r, column=4).number_format = currency_format
    for c in range(1, 5):
        ws_dash.cell(row=r, column=c).border = thin_border
        if i % 2 == 0:
            ws_dash.cell(row=r, column=c).fill = alt_fill

total_cat_row = cat_h + 1 + len(CATEGORIES)
ws_dash.cell(row=total_cat_row, column=1, value="TOTAL").font = Font(bold=True)
ws_dash.cell(
    row=total_cat_row, column=2
).value = f"=SUM(B{cat_h + 1}:B{total_cat_row - 1})"
ws_dash.cell(row=total_cat_row, column=2).number_format = currency_format
ws_dash.cell(row=total_cat_row, column=2).font = Font(bold=True)
ws_dash.cell(row=total_cat_row, column=3).value = 1
ws_dash.cell(row=total_cat_row, column=3).number_format = percent_format
for c in range(1, 5):
    ws_dash.cell(row=total_cat_row, column=c).border = thin_border
    ws_dash.cell(row=total_cat_row, column=c).fill = PatternFill(
        start_color=GRAY, end_color=GRAY, fill_type="solid"
    )

# ── Savings Overview ──
sav_start = total_cat_row + 2
section_title(ws_dash, sav_start, 1, 4, "SAVINGS OVERVIEW", GREEN)

sav_h = sav_start + 1
for col, h in enumerate(
    ["Account", "Total Deposits", "Total Withdrawals", "Net Savings"], 1
):
    ws_dash.cell(row=sav_h, column=col, value=h)
style_header_row(ws_dash, sav_h, 4)

sav_alt = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
for i, acct in enumerate(SAVINGS_ACCOUNTS):
    r = sav_h + 1 + i
    ws_dash.cell(row=r, column=1, value=acct)
    ws_dash.cell(
        row=r, column=2
    ).value = f'=SUMPRODUCT((EXACT(Savings!B3:B502,A{r}))*(EXACT(Savings!C3:C502,"Deposit"))*(Savings!D3:D502))'
    ws_dash.cell(row=r, column=2).number_format = currency_format
    ws_dash.cell(
        row=r, column=3
    ).value = f'=SUMPRODUCT((EXACT(Savings!B3:B502,A{r}))*(EXACT(Savings!C3:C502,"Withdrawal"))*(Savings!D3:D502))'
    ws_dash.cell(row=r, column=3).number_format = currency_format
    ws_dash.cell(row=r, column=4).value = f"=B{r}-C{r}"
    ws_dash.cell(row=r, column=4).number_format = currency_format
    for c in range(1, 5):
        ws_dash.cell(row=r, column=c).border = thin_border
        if i % 2 == 0:
            ws_dash.cell(row=r, column=c).fill = sav_alt

sav_total = sav_h + 1 + len(SAVINGS_ACCOUNTS)
ws_dash.cell(row=sav_total, column=1, value="TOTAL").font = Font(bold=True)
for c_idx in [2, 3, 4]:
    ws_dash.cell(
        row=sav_total, column=c_idx
    ).value = f"=SUM({get_column_letter(c_idx)}{sav_h + 1}:{get_column_letter(c_idx)}{sav_total - 1})"
    ws_dash.cell(row=sav_total, column=c_idx).number_format = currency_format
ws_dash.cell(row=sav_total, column=4).font = Font(bold=True)
for c in range(1, 5):
    ws_dash.cell(row=sav_total, column=c).border = thin_border
    ws_dash.cell(row=sav_total, column=c).fill = PatternFill(
        start_color=GRAY, end_color=GRAY, fill_type="solid"
    )

# ── Monthly Breakdown (cols F-I) ──
monthly_start = 3
section_title(ws_dash, monthly_start, 6, 9, "MONTHLY BREAKDOWN", BLUE)

m_h = monthly_start + 1
for col, h in enumerate(["Month", "Income", "Spending", "Net"], 6):
    ws_dash.cell(row=m_h, column=col, value=h)
style_header_row(ws_dash, m_h, 9)

months = [f"2026-{str(m).zfill(2)}" for m in range(1, 13)]
for i, month in enumerate(months):
    r = m_h + 1 + i
    ws_dash.cell(row=r, column=6, value=month)
    ws_dash.cell(
        row=r, column=7
    ).value = f"=SUMPRODUCT((EXACT(Income!F3:F502,F{r}))*(Income!D3:D502))"
    ws_dash.cell(row=r, column=7).number_format = currency_format
    ws_dash.cell(
        row=r, column=8
    ).value = (
        f"=SUMPRODUCT((EXACT(Transactions!G3:G1002,F{r}))*(Transactions!E3:E1002))"
    )
    ws_dash.cell(row=r, column=8).number_format = currency_format
    ws_dash.cell(row=r, column=9).value = f"=G{r}-H{r}"
    ws_dash.cell(row=r, column=9).number_format = currency_format
    for c in range(6, 10):
        ws_dash.cell(row=r, column=c).border = thin_border
        if i % 2 == 0:
            ws_dash.cell(row=r, column=c).fill = alt_fill

yr_total = m_h + 1 + len(months)
ws_dash.cell(row=yr_total, column=6, value="YEARLY TOTAL").font = Font(bold=True)
for c_idx in [7, 8, 9]:
    ws_dash.cell(
        row=yr_total, column=c_idx
    ).value = f"=SUM({get_column_letter(c_idx)}{m_h + 1}:{get_column_letter(c_idx)}{yr_total - 1})"
    ws_dash.cell(row=yr_total, column=c_idx).number_format = currency_format
ws_dash.cell(row=yr_total, column=9).font = Font(bold=True)
for c in range(6, 10):
    ws_dash.cell(row=yr_total, column=c).border = thin_border
    ws_dash.cell(row=yr_total, column=c).fill = PatternFill(
        start_color=GRAY, end_color=GRAY, fill_type="solid"
    )

# ── Column widths for Dashboard ──
for col, w in {1: 22, 2: 18, 3: 18, 4: 18, 5: 3, 6: 14, 7: 18, 8: 18, 9: 18}.items():
    ws_dash.column_dimensions[get_column_letter(col)].width = w

# ── Charts ──
pie = PieChart()
pie.title = "Spending by Category"
pie.style = 10
pie.add_data(
    Reference(ws_dash, min_col=2, min_row=cat_h + 1, max_row=cat_h + len(CATEGORIES))
)
pie.set_categories(
    Reference(ws_dash, min_col=1, min_row=cat_h + 1, max_row=cat_h + len(CATEGORIES))
)
pie.width = 18
pie.height = 13
ws_dash.add_chart(pie, f"F{sav_start}")

bar = BarChart()
bar.type = "col"
bar.title = "Monthly Income vs Spending"
bar.style = 10
bar.y_axis.title = "Amount"
bar.add_data(
    Reference(ws_dash, min_col=7, min_row=m_h, max_row=m_h + len(months)),
    titles_from_data=True,
)
bar.add_data(
    Reference(ws_dash, min_col=8, min_row=m_h, max_row=m_h + len(months)),
    titles_from_data=True,
)
bar.set_categories(
    Reference(ws_dash, min_col=6, min_row=m_h + 1, max_row=m_h + len(months))
)
bar.width = 22
bar.height = 13
ws_dash.add_chart(bar, f"F{sav_start + 20}")


# ═══════════════════════════════════════════════════════════════
# SHEET 7: FINANCIAL ADVISOR (Indonesian Smart Analysis)
# ═══════════════════════════════════════════════════════════════
ws_adv = wb.create_sheet("Advisor")
ws_adv.sheet_properties.tabColor = DARK_GREEN

ws_adv.merge_cells("A1:I1")
ws_adv["A1"].value = "PENASIHAT KEUANGAN / FINANCIAL ADVISOR"
ws_adv["A1"].font = Font(name="Calibri", size=18, bold=True, color=DARK_GREEN)
ws_adv["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_adv.row_dimensions[1].height = 45
ws_adv.merge_cells("A2:I2")
ws_adv[
    "A2"
].value = "Analisis otomatis berdasarkan standar keuangan Indonesia (OJK & BI)"
ws_adv["A2"].font = Font(name="Calibri", size=10, italic=True, color=DARK_GRAY)
ws_adv["A2"].alignment = Alignment(horizontal="center")

for col, w in {1: 32, 2: 20, 3: 20, 4: 22, 5: 3, 6: 30, 7: 22, 8: 22, 9: 55}.items():
    ws_adv.column_dimensions[get_column_letter(col)].width = w

good_fill = PatternFill(
    start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid"
)
warn_fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
bad_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")

# ──────────────────────────────────────
# LEFT: Indonesian Financial Health Scorecard
# ──────────────────────────────────────
row = 4
section_title(ws_adv, row, 1, 4, "SKOR KESEHATAN KEUANGAN", DARK_GREEN)

ratio_header = row + 1
for col, h in enumerate(
    ["Indikator / Indicator", "Nilai Anda", "Target (ID)", "Status"], 1
):
    ws_adv.cell(row=ratio_header, column=col, value=h)
style_header_row(ws_adv, ratio_header, 4)

ratios = [
    (
        "Rasio Tabungan (Savings Rate)",
        "=IF(Dashboard!B4>0,(Dashboard!B4-Dashboard!B5)/Dashboard!B4,0)",
        ">= 20% (ideal ID)",
        '=IF(B6>=0.2,"SEHAT",IF(B6>=0.1,"CUKUP",IF(B6>=0.05,"MINIMUM","KURANG")))',
        percent_format,
    ),
    (
        "Dana Darurat (Emergency Fund)",
        '=IF(Dashboard!B5>0,SUMPRODUCT((EXACT(Savings!B3:B502,"Emergency Fund"))*(EXACT(Savings!C3:C502,"Deposit"))*(Savings!D3:D502))/(Dashboard!B5/MAX(COUNTIF(Transactions!G3:G1002,"<>"),1)),0)',
        ">= 6 bulan (karyawan tetap)",
        '=IF(B7>=6,"SEHAT",IF(B7>=3,"CUKUP","KURANG - Prioritaskan!"))',
        "0.0",
    ),
    (
        "Rasio Pengeluaran (Spending)",
        "=IF(Dashboard!B4>0,Dashboard!B5/Dashboard!B4,0)",
        "<= 70% dari income",
        '=IF(B8<=0.7,"SEHAT",IF(B8<=0.85,"AWAS","BOROS - Kurangi!"))',
        percent_format,
    ),
    (
        "Kepatuhan Budget",
        '=IF(COUNTA(Budget!E3:E10)>0,COUNTIF(Budget!E3:E10,"OK")/COUNTA(Budget!E3:E10),0)',
        "100% on budget",
        '=IF(B9>=0.9,"SEHAT",IF(B9>=0.7,"CUKUP","KURANG - Review budget"))',
        percent_format,
    ),
    (
        "Kebutuhan vs Keinginan",
        '=IF(Dashboard!B5>0,(SUMPRODUCT((EXACT(Transactions!C3:C1002,"Food & Groceries"))*(Transactions!E3:E1002))+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Housing"))*(Transactions!E3:E1002))+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Bills & Utilities"))*(Transactions!E3:E1002))+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Healthcare"))*(Transactions!E3:E1002))+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Transportation"))*(Transactions!E3:E1002)))/Dashboard!B5,0)',
        "40-50% kebutuhan pokok",
        '=IF(AND(B10>=0.35,B10<=0.55),"SEHAT",IF(B10>0.55,"TERLALU TINGGI","RENDAH"))',
        percent_format,
    ),
    (
        "Alokasi Investasi (% NW)",
        "=IF(Dashboard!B9>0,Dashboard!B8/Dashboard!B9,0)",
        ">= 30% dari net worth",
        '=IF(B11>=0.3,"SEHAT",IF(B11>=0.15,"CUKUP","KURANG - Mulai investasi!"))',
        percent_format,
    ),
    (
        "Rasio Utang (Debt-to-Income)",
        "=0",
        "< 30% (BI standard)",
        '=IF(B12<0.3,"SEHAT",IF(B12<0.4,"AWAS","BAHAYA - BI limit 30-40%"))',
        percent_format,
    ),
    (
        "Pertumbuhan Net Worth",
        "=IF(Dashboard!B9>0,Dashboard!B6/Dashboard!B9,0)",
        "> inflasi (3.5%)",
        '=IF(B13>0.035,"SEHAT",IF(B13>0,"TUMBUH","MENURUN"))',
        percent_format,
    ),
]

for i, (name, formula, target, status_formula, fmt) in enumerate(ratios):
    r = ratio_header + 1 + i
    ws_adv.cell(row=r, column=1, value=name).font = Font(
        name="Calibri", size=11, bold=True
    )
    ws_adv.cell(row=r, column=2, value=formula).number_format = fmt
    ws_adv.cell(row=r, column=3, value=target).font = Font(
        name="Calibri", size=10, italic=True
    )
    ws_adv.cell(row=r, column=4, value=status_formula).font = Font(
        name="Calibri", size=11, bold=True
    )
    for c in range(1, 5):
        ws_adv.cell(row=r, column=c).border = thin_border

# ── Overall Score ──
score_row = ratio_header + 1 + len(ratios) + 1
ws_adv.cell(row=score_row, column=1, value="SKOR KESEHATAN TOTAL").font = Font(
    name="Calibri", size=13, bold=True, color=DARK_GREEN
)
ws_adv.cell(row=score_row, column=2).value = (
    '=ROUND((COUNTIF(D6:D13,"SEHAT")*3+COUNTIF(D6:D13,"CUKUP")*2'
    '+COUNTIF(D6:D13,"TUMBUH")*3+COUNTIF(D6:D13,"MINIMUM")*1'
    '+COUNTIF(D6:D13,"AWAS")*1)/(COUNTA(D6:D13)*3)*100,0)'
)
ws_adv.cell(row=score_row, column=2).number_format = '0" / 100"'
ws_adv.cell(row=score_row, column=2).font = Font(
    name="Calibri", size=16, bold=True, color=DARK_GREEN
)
ws_adv.cell(row=score_row, column=3).value = (
    f'=IF(B{score_row}>=80,"Sangat Baik",IF(B{score_row}>=60,"Baik",'
    f'IF(B{score_row}>=40,"Cukup","Perlu Perhatian")))'
)
ws_adv.cell(row=score_row, column=3).font = Font(name="Calibri", size=13, bold=True)
for c in range(1, 5):
    ws_adv.cell(row=score_row, column=c).border = Border(
        left=Side(style="medium", color=DARK_GREEN),
        right=Side(style="medium", color=DARK_GREEN),
        top=Side(style="medium", color=DARK_GREEN),
        bottom=Side(style="medium", color=DARK_GREEN),
    )
    ws_adv.cell(row=score_row, column=c).fill = good_fill


# ──────────────────────────────────────
# LEFT: Indonesian Adapted Budget Rule (40/30/20/10)
# ──────────────────────────────────────
rule_start = score_row + 2
section_title(
    ws_adv, rule_start, 1, 4, "ATURAN 40/30/20/10 (Adaptasi Indonesia)", DARK_GREEN
)

rule_h = rule_start + 1
for col, h in enumerate(["Kategori", "Ideal %", "Aktual %", "Jumlah (Rp)"], 1):
    ws_adv.cell(row=rule_h, column=col, value=h)
style_header_row(ws_adv, rule_h, 4)

# Indonesian adapted formula: 40% needs, 30% wants, 20% savings, 10% charity/family
needs_formula = (
    "=IF(Dashboard!B4>0,"
    '(SUMPRODUCT((EXACT(Transactions!C3:C1002,"Food & Groceries"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Housing"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Transportation"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Healthcare"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Bills & Utilities"))*(Transactions!E3:E1002)))'
    "/Dashboard!B4,0)"
)
needs_amt = (
    '=SUMPRODUCT((EXACT(Transactions!C3:C1002,"Food & Groceries"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Housing"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Transportation"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Healthcare"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Bills & Utilities"))*(Transactions!E3:E1002))'
)
wants_formula = (
    "=IF(Dashboard!B4>0,"
    '(SUMPRODUCT((EXACT(Transactions!C3:C1002,"Entertainment"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Shopping"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Education"))*(Transactions!E3:E1002)))'
    "/Dashboard!B4,0)"
)
wants_amt = (
    '=SUMPRODUCT((EXACT(Transactions!C3:C1002,"Entertainment"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Shopping"))*(Transactions!E3:E1002))'
    '+SUMPRODUCT((EXACT(Transactions!C3:C1002,"Education"))*(Transactions!E3:E1002))'
)
savings_pct = "=IF(Dashboard!B4>0,(Dashboard!B4-Dashboard!B5)/Dashboard!B4,0)"
savings_amt = "=Dashboard!B4-Dashboard!B5"

rule_items = [
    ("Kebutuhan Pokok (Living)", 0.40, needs_formula, needs_amt),
    ("Keinginan (Lifestyle)", 0.30, wants_formula, wants_amt),
    ("Tabungan & Investasi", 0.20, savings_pct, savings_amt),
    ("Sedekah/Zakat/Keluarga", 0.10, "=0", "=0"),
]
for i, (label, ideal, your_pct, your_amt) in enumerate(rule_items):
    r = rule_h + 1 + i
    ws_adv.cell(row=r, column=1, value=label).font = Font(
        name="Calibri", size=11, bold=True
    )
    ws_adv.cell(row=r, column=2, value=ideal).number_format = percent_format
    ws_adv.cell(row=r, column=3, value=your_pct).number_format = percent_format
    ws_adv.cell(row=r, column=4, value=your_amt).number_format = currency_format
    for c in range(1, 5):
        ws_adv.cell(row=r, column=c).border = thin_border

verdict_row = rule_h + 1 + len(rule_items)
ws_adv.cell(row=verdict_row, column=1, value="Penilaian").font = Font(bold=True)
ws_adv.cell(
    row=verdict_row, column=2
).value = '=IF(AND(C{n}<=0.50,C{w}<=0.35,C{s}>=0.15),"Seimbang / Balanced","Perlu Penyesuaian")'.format(
    n=rule_h + 1, w=rule_h + 2, s=rule_h + 3
)
ws_adv.cell(row=verdict_row, column=2).font = Font(size=12, bold=True)
for c in range(1, 5):
    ws_adv.cell(row=verdict_row, column=c).border = thin_border
    ws_adv.cell(row=verdict_row, column=c).fill = warn_fill


# ──────────────────────────────────────
# LEFT: Indonesian Investment Guide
# ──────────────────────────────────────
inv_start = verdict_row + 2
section_title(ws_adv, inv_start, 1, 4, "PANDUAN INVESTASI INDONESIA", PURPLE)

inv_h = inv_start + 1
for col, h in enumerate(["Instrumen", "Return/thn", "Risiko", "Pajak"], 1):
    ws_adv.cell(row=inv_h, column=col, value=h)
style_header_row(ws_adv, inv_h, 4)

investments = [
    ("Tabungan Bank", "0.5-2%", "Sangat Rendah", "20% bunga >Rp7.5jt"),
    ("Deposito", "3-5%", "Rendah", "20% bunga"),
    ("Reksadana Pasar Uang", "3-5%", "Rendah", "BEBAS PAJAK"),
    ("Reksadana Pendapatan Tetap", "5-8%", "Rendah-Sedang", "BEBAS PAJAK"),
    ("Reksadana Campuran", "6-12%", "Sedang", "BEBAS PAJAK"),
    ("Reksadana Saham", "8-15%", "Tinggi", "BEBAS PAJAK"),
    ("SBN Ritel (ORI/SBR/SR/ST)", "5-7%", "Rendah (pemerintah)", "10% kupon"),
    ("Saham (IDX)", "10-15%", "Tinggi", "0.1% jual"),
    ("Emas (Antam/digital)", "5-10%", "Sedang", "PPh capital gain"),
    ("P2P Lending (OJK)", "10-18%", "Tinggi-Sangat Tinggi", "15% bunga"),
]
for i, (name, ret, risk, tax) in enumerate(investments):
    r = inv_h + 1 + i
    ws_adv.cell(row=r, column=1, value=name).font = Font(name="Calibri", size=10)
    ws_adv.cell(row=r, column=2, value=ret).font = Font(name="Calibri", size=10)
    ws_adv.cell(row=r, column=3, value=risk).font = Font(name="Calibri", size=10)
    ws_adv.cell(row=r, column=4, value=tax).font = Font(name="Calibri", size=10)
    for c in range(1, 5):
        ws_adv.cell(row=r, column=c).border = thin_border
        if i % 2 == 0:
            ws_adv.cell(row=r, column=c).fill = PatternFill(
                start_color=LIGHT_PURPLE, end_color=LIGHT_PURPLE, fill_type="solid"
            )
    # Highlight tax-free
    if "BEBAS" in tax:
        ws_adv.cell(row=r, column=4).font = Font(
            name="Calibri", size=10, bold=True, color=GREEN
        )


# ──────────────────────────────────────
# RIGHT: Monthly Trend Analysis (Indonesian advice)
# ──────────────────────────────────────
section_title(ws_adv, 4, 6, 9, "ANALISIS TREN BULANAN", DARK_GREEN)

trend_h = 5
for col, h in enumerate(
    ["Bulan", "Savings Rate", "Pengeluaran", "Nasihat / Advice"], 6
):
    ws_adv.cell(row=trend_h, column=col, value=h)
style_header_row(ws_adv, trend_h, 9)

for i, month in enumerate(months):
    r = trend_h + 1 + i
    ws_adv.cell(row=r, column=6, value=month)

    inc_ref = f"SUMPRODUCT((EXACT(Income!F3:F502,F{r}))*(Income!D3:D502))"
    exp_ref = f"SUMPRODUCT((EXACT(Transactions!G3:G1002,F{r}))*(Transactions!E3:E1002))"

    # Savings rate
    ws_adv.cell(
        row=r, column=7
    ).value = f'=IF({inc_ref}>0,({inc_ref}-{exp_ref})/{inc_ref},"")'
    ws_adv.cell(row=r, column=7).number_format = percent_format

    # Spending total
    ws_adv.cell(row=r, column=8).value = f"={exp_ref}"
    ws_adv.cell(row=r, column=8).number_format = currency_format

    # Indonesian monthly advice
    ws_adv.cell(row=r, column=9).value = (
        f'=IF({inc_ref}=0,"Belum ada data",'
        f"IF(({inc_ref}-{exp_ref})/{inc_ref}>=0.3,"
        f'"Luar biasa! Nabung >30%. Pertimbangkan investasi reksadana/SBN.",'
        f"IF(({inc_ref}-{exp_ref})/{inc_ref}>=0.2,"
        f'"Bagus. Sesuai target 20%. Tetap konsisten!",'
        f"IF(({inc_ref}-{exp_ref})/{inc_ref}>=0.1,"
        f'"Cukup, tapi di bawah ideal. Kurangi jajan & belanja online.",'
        f"IF(({inc_ref}-{exp_ref})/{inc_ref}>=0,"
        f'"Peringatan: Nabung <10%. Cek tagihan & langganan.",'
        f'"BAHAYA: Pengeluaran > pemasukan! Segera kurangi pengeluaran.")))))'
    )
    ws_adv.cell(row=r, column=9).font = Font(name="Calibri", size=10)
    ws_adv.cell(row=r, column=9).alignment = Alignment(wrap_text=True)
    ws_adv.row_dimensions[r].height = 30

    for c in range(6, 10):
        ws_adv.cell(row=r, column=c).border = thin_border
        if i % 2 == 0:
            ws_adv.cell(row=r, column=c).fill = good_fill


# ──────────────────────────────────────
# RIGHT: Smart Alerts (Indonesian context)
# ──────────────────────────────────────
alert_start = trend_h + 1 + len(months) + 1
section_title(ws_adv, alert_start, 6, 9, "PERINGATAN CERDAS / SMART ALERTS", RED)

alert_h = alert_start + 1
for col, h in enumerate(["Peringatan", "Detail", "Nilai", "Saran"], 6):
    ws_adv.cell(row=alert_h, column=col, value=h)
style_header_row(ws_adv, alert_h, 9)

smart_alerts = [
    (
        "Kategori Tertinggi",
        "=INDEX(Dashboard!A13:A20,MATCH(MAX(Dashboard!B13:B20),Dashboard!B13:B20,0))",
        '=TEXT(MAX(Dashboard!B13:B20),"#,##0")',
        "Pastikan sesuai prioritas kebutuhan",
    ),
    (
        "Over Budget",
        '=COUNTIF(Budget!E3:E10,"OVER BUDGET")&" kategori melebihi budget"',
        '=IF(COUNTIF(Budget!E3:E10,"OVER BUDGET")>0,"Perlu tindakan","Semua on track!")',
        '=IF(COUNTIF(Budget!E3:E10,"OVER BUDGET")>2,"Serius: banyak kategori over budget!","Cek ulang budget bulanan")',
    ),
    (
        "Transaksi Terbesar",
        "=MAX(Transactions!E3:E1002)",
        '=IF(MAX(Transactions!E3:E1002)>Dashboard!B4*0.1,">10% income","Normal")',
        '=IF(MAX(Transactions!E3:E1002)>Dashboard!B4*0.1,"Pengeluaran besar. Pertimbangkan cicilan atau dana darurat.","Dalam batas wajar")',
    ),
    (
        "Rasio Makan (%)",
        '=IF(Dashboard!B5>0,SUMPRODUCT((EXACT(Transactions!C3:C1002,"Food & Groceries"))*(Transactions!E3:E1002))/Dashboard!B5,0)',
        "",
        '=IF(H{r}>0.3,"Makan >30% pengeluaran. Coba meal prep & masak sendiri.","Porsi makan wajar")',
    ),
    (
        "Progress Tabungan",
        "=IF(SUMPRODUCT((Savings!F3:F502)*1)>0,Dashboard!B7/SUMPRODUCT((Savings!F3:F502)*1),0)",
        "",
        '=IF(H{r}>=0.5,"Sudah >50% target! Terus nabung!",IF(H{r}>0,"Progress "&TEXT(H{r},"0%")&" - terus konsisten!","Tetapkan target tabungan"))',
    ),
    (
        "Dana Darurat Status",
        '=IF(Dashboard!B5>0,SUMPRODUCT((EXACT(Savings!B3:B502,"Emergency Fund"))*(EXACT(Savings!C3:C502,"Deposit"))*(Savings!D3:D502))/(Dashboard!B5/MAX(COUNTIF(Transactions!G3:G1002,"<>"),1)),0)',
        "",
        '=IF(H{r}>=6,"Dana darurat aman (>6 bulan).",IF(H{r}>=3,"Dana darurat cukup (3-6 bulan). Tambah terus.","PRIORITAS: Dana darurat kurang dari 3 bulan!"))',
    ),
    (
        "Kesiapan Zakat",
        '=IF(Dashboard!B9>=Config!B9,"Harta di atas nisab. Siapkan zakat 2.5%.","Belum wajib zakat mal")',
        "=IF(Dashboard!B9>=Config!B9,Dashboard!B9*0.025,0)",
        '=IF(Dashboard!B9>=Config!B9,"Perkiraan zakat: Rp "&TEXT(Dashboard!B9*0.025,"#,##0")&". Bayar via BAZNAS (bisa jadi pengurang pajak).","")',
    ),
]

for i, (label, detail_f, value_f, action_f) in enumerate(smart_alerts):
    r = alert_h + 1 + i
    ws_adv.cell(row=r, column=6, value=label).font = Font(
        name="Calibri", size=11, bold=True
    )
    ws_adv.cell(row=r, column=7, value=detail_f)
    ws_adv.cell(
        row=r,
        column=8,
        value=value_f.replace("{r}", str(r)) if "{r}" in value_f else value_f,
    )
    ws_adv.cell(
        row=r,
        column=9,
        value=action_f.replace("{r}", str(r)) if "{r}" in action_f else action_f,
    )

    if i in (3, 4):
        ws_adv.cell(row=r, column=7).number_format = percent_format
    elif i == 2:
        ws_adv.cell(row=r, column=7).number_format = currency_format
    elif i == 5:
        ws_adv.cell(row=r, column=7).number_format = "0.0"
    if i == 6:
        ws_adv.cell(row=r, column=8).number_format = currency_format

    ws_adv.cell(row=r, column=9).font = Font(name="Calibri", size=10)
    ws_adv.cell(row=r, column=9).alignment = Alignment(wrap_text=True)
    ws_adv.row_dimensions[r].height = 32
    for c in range(6, 10):
        ws_adv.cell(row=r, column=c).border = thin_border
        ws_adv.cell(row=r, column=c).fill = bad_fill


# ──────────────────────────────────────
# RIGHT: Indonesian Financial Tips
# ──────────────────────────────────────
tips_start = alert_h + 1 + len(smart_alerts) + 2
section_title(ws_adv, tips_start, 6, 9, "TIPS KEUANGAN INDONESIA", PURPLE)

tips = [
    "Bayar diri sendiri dulu: Sisihkan tabungan & investasi DI AWAL gajian, bukan sisa.",
    "Aturan 40/30/20/10: 40% kebutuhan, 30% keinginan, 20% tabungan/investasi, 10% sedekah/keluarga.",
    "Dana darurat 6 bulan: Simpan di reksadana pasar uang (return 3-5%, bebas pajak, mudah dicairkan).",
    "Investasi mulai Rp 10rb: Gunakan reksadana (Bibit, Bareksa, IPOT) - bebas pajak untuk individu.",
    "SBN Ritel (ORI/SBR): Obligasi pemerintah, dijamin negara, return 5-7%, pajak hanya 10%.",
    "Hindari pinjol ilegal: Cek daftar OJK. Bunga pinjol legal max 0.4%/hari sudah sangat mahal.",
    "Zakat sebagai pengurang pajak: Bayar via BAZNAS, lampirkan bukti di SPT (mengurangi PKP).",
    "BPJS JHT = tabungan jangka panjang: Jangan dicairkan dini. Tambah dengan DPLK (pengurang pajak 5%).",
    "Aturan 24 jam: Untuk pembelian > Rp 200rb yang bukan kebutuhan, tunggu 24 jam sebelum beli.",
    "THR & bonus: Gunakan rumus 50% tabungan/investasi, 30% kebutuhan, 20% keinginan.",
    "Catat uang kiriman keluarga: Budget kirim orang tua sebagai pengeluaran tetap (10-20% gaji).",
    "Arisan = tabungan sosial: Catat sebagai pengeluaran tetap. Jangan ambil arisan untuk konsumsi.",
]
for i, tip in enumerate(tips):
    r = tips_start + 1 + i
    ws_adv.cell(row=r, column=6, value=f"{i + 1}.").font = Font(bold=True, color=PURPLE)
    ws_adv.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
    ws_adv.cell(row=r, column=7, value=tip).font = Font(name="Calibri", size=10)
    ws_adv.cell(row=r, column=7).alignment = Alignment(wrap_text=True)
    ws_adv.row_dimensions[r].height = 30
    for c in range(6, 10):
        ws_adv.cell(row=r, column=c).border = thin_border
        if i % 2 == 0:
            ws_adv.cell(row=r, column=c).fill = PatternFill(
                start_color=LIGHT_PURPLE, end_color=LIGHT_PURPLE, fill_type="solid"
            )


# ──────────────────────────────────────
# LEFT BOTTOM: Current Status Summary (Indonesian)
# ──────────────────────────────────────
status_start = inv_start + len(investments) + 3
section_title(ws_adv, status_start, 1, 4, "RINGKASAN STATUS KEUANGAN", DARK_GREEN)

status_items = [
    (
        "Pendapatan Bulanan (avg)",
        '=IF(COUNTIF(Income!F3:F502,"<>")>0,SUMPRODUCT((Income!D3:D502)*1)/MAX(COUNTIF(Income!F3:F502,"<>"),1),0)',
    ),
    (
        "Pengeluaran Bulanan (avg)",
        '=IF(COUNTIF(Transactions!G3:G1002,"<>")>0,SUMPRODUCT((Transactions!E3:E1002)*1)/MAX(COUNTIF(Transactions!G3:G1002,"<>"),1),0)',
    ),
    ("Surplus/Defisit Bulanan", f"=B{status_start + 2}-B{status_start + 3}"),
    ("Total Tabungan Likuid", "=Dashboard!B7"),
    ("Total Nilai Aset", "=Dashboard!B8"),
    ("Kekayaan Bersih (Net Worth)", "=Dashboard!B9"),
    (
        "Runway (bulan bertahan)",
        f"=IF(B{status_start + 3}>0,B{status_start + 5}/B{status_start + 3},0)",
    ),
    ("Estimasi Zakat Mal/tahun", f"=IF(Dashboard!B9>=Config!B9,Dashboard!B9*0.025,0)"),
    (
        "PPh 21 Estimasi/tahun",
        f"=IF(B{status_start + 2}*12>Config!B3,IF(B{status_start + 2}*12-Config!B3<=60000000,(B{status_start + 2}*12-Config!B3)*0.05,(B{status_start + 2}*12-Config!B3)*0.15),0)",
    ),
]

for i, (label, formula) in enumerate(status_items):
    r = status_start + 2 + i
    ws_adv.cell(row=r, column=1, value=label).font = Font(
        name="Calibri", size=11, bold=True
    )
    ws_adv.cell(row=r, column=1).fill = subheader_fill
    ws_adv.cell(row=r, column=2, value=formula)
    if "bulan" in label.lower() or "runway" in label.lower():
        ws_adv.cell(row=r, column=2).number_format = "0.0"
    else:
        ws_adv.cell(row=r, column=2).number_format = currency_format
    ws_adv.cell(row=r, column=2).font = Font(name="Calibri", size=12, bold=True)
    for c in range(1, 5):
        ws_adv.cell(row=r, column=c).border = thin_border


# ──────────────────────────────────────
# LEFT BOTTOM: BPJS & Tax Reference
# ──────────────────────────────────────
bpjs_start = status_start + 2 + len(status_items) + 1
section_title(ws_adv, bpjs_start, 1, 4, "REFERENSI BPJS & PAJAK", ORANGE)

bpjs_items = [
    ("BPJS Kesehatan Kelas 1", "Rp 150,000/bulan", "Wajib", "Perlindungan kesehatan"),
    ("BPJS Kesehatan Kelas 2", "Rp 100,000/bulan", "Wajib", "Perlindungan kesehatan"),
    ("BPJS Kesehatan Kelas 3", "Rp 35,000/bulan", "Wajib", "Perlindungan kesehatan"),
    (
        "BPJS JHT (karyawan)",
        "2% gaji (Anda) + 3.7% (perusahaan)",
        "Wajib",
        "Tabungan pensiun, cair usia 56",
    ),
    (
        "BPJS JP (karyawan)",
        "1% gaji (Anda) + 2% (perusahaan)",
        "Wajib",
        "Pensiun bulanan usia 56",
    ),
    (
        "DPLK (sukarela)",
        "Max 5% gaji = pengurang pajak",
        "Opsional",
        "Tambahan pensiun, hemat pajak",
    ),
    ("PPh 21 Bracket 1", "5% (s/d Rp 60jt/thn)", "Otomatis", "PTKP TK/0: Rp 54jt"),
    ("PPh 21 Bracket 2", "15% (Rp 60-250jt/thn)", "Otomatis", "PTKP K/0: Rp 58.5jt"),
]
for i, (item, detail, status, note) in enumerate(bpjs_items):
    r = bpjs_start + 1 + i
    ws_adv.cell(row=r, column=1, value=item).font = Font(name="Calibri", size=10)
    ws_adv.cell(row=r, column=2, value=detail).font = Font(name="Calibri", size=10)
    ws_adv.cell(row=r, column=3, value=status).font = Font(name="Calibri", size=10)
    ws_adv.cell(row=r, column=4, value=note).font = Font(name="Calibri", size=10)
    for c in range(1, 5):
        ws_adv.cell(row=r, column=c).border = thin_border
        if i % 2 == 0:
            ws_adv.cell(row=r, column=c).fill = PatternFill(
                start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid"
            )


# ═══════════════════════════════════════════════════════════════
# SHEET 8: CONFIG (Indonesian Financial Constants)
# ═══════════════════════════════════════════════════════════════
ws_cfg = wb.create_sheet("Config")
ws_cfg.sheet_properties.tabColor = DARK_GRAY

ws_cfg.merge_cells("A1:C1")
ws_cfg["A1"].value = "KONFIGURASI KEUANGAN INDONESIA"
ws_cfg["A1"].font = title_font
ws_cfg["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_cfg.row_dimensions[1].height = 35

for col, h in enumerate(["Parameter", "Nilai", "Keterangan"], 1):
    ws_cfg.cell(row=2, column=col, value=h)
style_header_row(ws_cfg, 2, 3)

ws_cfg.column_dimensions["A"].width = 35
ws_cfg.column_dimensions["B"].width = 22
ws_cfg.column_dimensions["C"].width = 45

config_data = [
    ("PTKP TK/0 (single)", 54000000, "Penghasilan Tidak Kena Pajak - Lajang"),
    ("PTKP K/0 (married)", 58500000, "Penghasilan Tidak Kena Pajak - Menikah"),
    ("PTKP K/1", 63000000, "Menikah + 1 tanggungan"),
    ("PTKP K/2", 67500000, "Menikah + 2 tanggungan"),
    ("PTKP K/3", 72000000, "Menikah + 3 tanggungan"),
    (
        "Nisab Zakat Mal (Rp)",
        85000000,
        "~85 gram emas x harga emas terkini. Update manual.",
    ),
    ("Tarif Zakat", 0.025, "2.5% dari harta di atas nisab"),
    ("Inflasi Target (BI)", 0.035, "Target Bank Indonesia 2.5% +/- 1%"),
    ("BI Rate", 0.0575, "Suku bunga acuan Bank Indonesia"),
    ("Pajak Deposito", 0.20, "PPh final atas bunga deposito"),
    ("Pajak SBN Ritel", 0.10, "PPh final atas kupon SBN"),
    ("Pajak Reksadana", 0, "Bebas pajak untuk individu"),
    ("Pajak Jual Saham", 0.001, "0.1% final atas nilai jual"),
    ("Debt-to-Income Max (BI)", 0.30, "Batas rasio utang Bank Indonesia"),
    ("BPJS JHT (karyawan)", 0.02, "2% dari gaji karyawan"),
    ("BPJS JP (karyawan)", 0.01, "1% dari gaji karyawan"),
    ("BPJS Kesehatan Kelas 1", 150000, "Per bulan per orang"),
    ("UMP Jakarta 2026", 5500000, "Upah Minimum Provinsi DKI Jakarta (est)"),
]

for i, (param, value, desc) in enumerate(config_data):
    r = 3 + i
    ws_cfg.cell(row=r, column=1, value=param).font = Font(name="Calibri", size=11)
    ws_cfg.cell(row=r, column=2, value=value)
    if isinstance(value, float) and value < 1:
        ws_cfg.cell(row=r, column=2).number_format = "0.00%"
    else:
        ws_cfg.cell(row=r, column=2).number_format = currency_format
    ws_cfg.cell(row=r, column=2).font = Font(name="Calibri", size=11, bold=True)
    ws_cfg.cell(row=r, column=3, value=desc).font = Font(
        name="Calibri", size=10, italic=True
    )
    for c in range(1, 4):
        ws_cfg.cell(row=r, column=c).border = thin_border
        if i % 2 == 0:
            ws_cfg.cell(row=r, column=c).fill = PatternFill(
                start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"
            )

# Note at bottom of config
note_row = 3 + len(config_data) + 1
ws_cfg.merge_cells(f"A{note_row}:C{note_row}")
ws_cfg.cell(row=note_row, column=1).value = (
    "Catatan: Nilai di atas bisa di-update sesuai kondisi terkini. "
    "Nisab zakat perlu disesuaikan dengan harga emas terbaru. "
    "PTKP dan PPh mengikuti UU HPP."
)
ws_cfg.cell(row=note_row, column=1).font = Font(
    name="Calibri", size=10, italic=True, color=RED
)
ws_cfg.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
ws_cfg.row_dimensions[note_row].height = 40


# ═══════════════════════════════════════════════════════════════
# SHEET ORDER: Dashboard first, then Advisor
# ═══════════════════════════════════════════════════════════════
# Current: Transactions(0), Income(1), Savings(2), Assets(3), Budget(4), Dashboard(5), Advisor(6), Config(7)
# Want:    Dashboard(0), Advisor(1), Transactions, Income, Savings, Assets, Budget, Config
wb.move_sheet("Dashboard", offset=-5)
wb.move_sheet("Advisor", offset=-5)

# ═══════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════
import os

filepath = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "Financial_Tracker.xlsx"
)
wb.save(filepath)
print(f"Financial tracker saved to: {filepath}")
print(
    "Sheets: Dashboard, Advisor, Transactions, Income, Savings, Assets, Budget, Config"
)
